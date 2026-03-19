"""Users Router - Profile, CRUD, Import, Batch, Documents"""
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional, List, Dict, Any
from datetime import datetime, timezone, timedelta
import io
import secrets

from routers import (
    db, get_current_user, generate_uuid, get_password_hash, verify_password,
    is_admin_user, logger,
)

router = APIRouter(prefix="/api", tags=["users"])


# --- Models ---

class UserUpdate(BaseModel):
    full_name: Optional[str] = None
    surname: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    department: Optional[str] = None
    division: Optional[str] = None
    position: Optional[str] = None
    bio: Optional[str] = None
    avatar_url: Optional[str] = None
    roles: Optional[List[str]] = None
    is_active: Optional[bool] = None
    student_id: Optional[str] = None
    team_id: Optional[str] = None


# --- User CRUD ---

@router.get("/users")
async def get_users(
    role: Optional[str] = None,
    status: Optional[str] = None,
    department: Optional[str] = None,
    search: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if status == "active":
        query["is_active"] = True
    elif status == "inactive":
        query["is_active"] = False
    if department:
        query["department"] = department
    if role:
        query["roles"] = role
    if search:
        query["$or"] = [
            {"full_name": {"$regex": search, "$options": "i"}},
            {"email": {"$regex": search, "$options": "i"}},
        ]
    users = await db.users.find(query, {"_id": 0, "password_hash": 0}).to_list(1000)
    return users


@router.get("/users/me")
async def get_my_profile(current_user: dict = Depends(get_current_user)):
    user = await db.users.find_one({"id": current_user["id"]}, {"_id": 0, "password_hash": 0})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    return user


@router.put("/users/me")
async def update_my_profile(update_data: dict, current_user: dict = Depends(get_current_user)):
    allowed = {"full_name", "phone", "department", "bio", "avatar_initials"}
    filtered = {k: v for k, v in update_data.items() if k in allowed and v is not None}
    filtered["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.users.update_one({"id": current_user["id"]}, {"$set": filtered})
    user = await db.users.find_one({"id": current_user["id"]}, {"_id": 0, "password_hash": 0})
    return user


@router.post("/users/me/change-password")
async def change_my_password(data: dict, current_user: dict = Depends(get_current_user)):
    current_password = data.get("current_password", "")
    new_password = data.get("new_password", "")
    if len(new_password) < 8:
        raise HTTPException(status_code=400, detail="New password must be at least 8 characters")
    user_record = await db.users.find_one({"id": current_user["id"]}, {"_id": 0})
    if not user_record:
        raise HTTPException(status_code=404, detail="User not found")
    if user_record.get("password_hash"):
        if not verify_password(current_password, user_record["password_hash"]):
            raise HTTPException(status_code=400, detail="Current password is incorrect")
    await db.users.update_one(
        {"id": current_user["id"]},
        {"$set": {"password_hash": get_password_hash(new_password), "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    return {"message": "Password updated successfully"}


@router.post("/users/invite")
async def invite_user(data: dict, current_user: dict = Depends(get_current_user)):
    if not is_admin_user(current_user):
        raise HTTPException(status_code=403, detail="Only admins can invite users")
    email = data.get("email", "").strip().lower()
    if not email:
        raise HTTPException(status_code=400, detail="Email is required")
    existing = await db.users.find_one({"email": email})
    if existing:
        raise HTTPException(status_code=409, detail="User with this email already exists")
    otp_code = "".join([str(secrets.randbelow(10)) for _ in range(6)])
    expires_at = datetime.now(timezone.utc) + timedelta(minutes=60)
    await db.otps.delete_many({"email": email})
    await db.otps.insert_one({"email": email, "otp": otp_code, "attempts": 0, "expires_at": expires_at, "created_at": datetime.now(timezone.utc)})
    return {"message": f"Invitation sent to {email}", "email": email, "role": data.get("role", "student"), "dev_note": f"Login OTP for invited user: {otp_code}"}


@router.post("/users")
async def create_user(data: dict, current_user: dict = Depends(get_current_user)):
    if not is_admin_user(current_user):
        raise HTTPException(status_code=403, detail="Only admins can create users")
    email = data.get("email", "").strip().lower()
    full_name = data.get("full_name", "")
    roles = data.get("roles", [data.get("role", "employee")])
    password = data.get("password", "")
    if not email:
        raise HTTPException(status_code=400, detail="Email is required")
    if not password or len(password) < 6:
        raise HTTPException(status_code=400, detail="Password must be at least 6 characters")
    existing = await db.users.find_one({"email": email})
    if existing:
        raise HTTPException(status_code=409, detail="User with this email already exists")
    user_dict = {
        "id": generate_uuid(),
        "email": email,
        "full_name": full_name or email.split("@")[0].title(),
        "roles": roles if isinstance(roles, list) else [roles],
        "is_verified": True, "is_active": True,
        "password_hash": get_password_hash(password),
        "division": data.get("division", ""), "department": data.get("department", ""),
        "position": data.get("position", ""),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.users.insert_one({**user_dict})
    return {k: v for k, v in user_dict.items() if k not in ["_id", "password_hash"]}


@router.get("/users/import-template")
async def download_import_template(current_user: dict = Depends(get_current_user)):
    if not is_admin_user(current_user):
        raise HTTPException(status_code=403, detail="Only admins can download import template")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    headers = [
        "email", "full_name", "surname", "personnel_number", "id_number",
        "gender", "race", "age", "division", "department", "position", "level",
        "start_date", "years_of_service",
        "ofo_major_group", "ofo_sub_major_group", "ofo_occupation", "ofo_code"
    ]
    sample_row = [
        "john.doe@sentech.co.za", "John Doe", "Doe", "EMP001", "9001015800088",
        "Male", "African", "34", "Engineering", "Software Dev", "Developer", "L5",
        "2020-01-15", "6",
        "PROFESSIONALS", "Science and Engineering", "Software Developer", "251201"
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "User Import Template"
    header_fill = PatternFill(start_color="0056B3", end_color="0056B3", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h.replace("_", " ").title())
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for col_idx, val in enumerate(sample_row, 1):
        ws.cell(row=2, column=col_idx, value=val)
    for col_idx, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(h) + 4, 15)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=user_import_template.xlsx"}
    )


@router.put("/users/{user_id}/status")
async def update_user_status(user_id: str, data: dict, current_user: dict = Depends(get_current_user)):
    if "admin" not in current_user.get("roles", []) and "manager" not in current_user.get("roles", []):
        raise HTTPException(status_code=403, detail="Access denied")
    is_active = data.get("is_active", True)
    await db.users.update_one({"id": user_id}, {"$set": {"is_active": is_active, "updated_at": datetime.now(timezone.utc).isoformat()}})
    return {"message": f"User {'activated' if is_active else 'deactivated'} successfully"}


@router.put("/users/{user_id}/role")
async def update_user_role(user_id: str, data: dict, current_user: dict = Depends(get_current_user)):
    if not is_admin_user(current_user):
        raise HTTPException(status_code=403, detail="Only admins can change roles")
    roles = data.get("roles", ["student"])
    valid_roles = {"admin", "student", "employee", "super_admin", "manager", "professional", "technician", "clerical"}
    roles = [r for r in roles if r in valid_roles]
    if not roles:
        roles = ["student"]
    await db.users.update_one({"id": user_id}, {"$set": {"roles": roles, "updated_at": datetime.now(timezone.utc).isoformat()}})
    return {"message": "User role updated successfully"}


@router.delete("/users/{user_id}")
async def delete_user(user_id: str, current_user: dict = Depends(get_current_user)):
    if not is_admin_user(current_user):
        raise HTTPException(status_code=403, detail="Only admins can delete users")
    if user_id == current_user["id"]:
        raise HTTPException(status_code=400, detail="Cannot delete your own account")
    result = await db.users.delete_one({"id": user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    return {"message": "User deleted successfully"}


@router.get("/users/{user_id}")
async def get_user(user_id: str, current_user: dict = Depends(get_current_user)):
    user = await db.users.find_one({"id": user_id}, {"_id": 0, "password_hash": 0})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    return user


@router.put("/users/{user_id}")
async def update_user(user_id: str, update_data: UserUpdate, current_user: dict = Depends(get_current_user)):
    admin = is_admin_user(current_user)
    if user_id != current_user["id"] and not admin:
        raise HTTPException(status_code=403, detail="Access denied")
    user = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    update_dict = {k: v for k, v in update_data.model_dump().items() if v is not None}
    update_dict["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.users.update_one({"id": user_id}, {"$set": update_dict})
    return {"message": "User updated successfully"}


# --- User Documents ---

@router.get("/users/{user_id}/documents")
async def get_user_documents(user_id: str, current_user: dict = Depends(get_current_user)):
    if user_id != current_user["id"] and not is_admin_user(current_user):
        raise HTTPException(status_code=403, detail="Access denied")
    user = await db.users.find_one({"id": user_id}, {"_id": 0, "documents": 1})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    return user.get("documents", [])


@router.post("/users/{user_id}/documents")
async def upload_user_document(user_id: str, data: dict, current_user: dict = Depends(get_current_user)):
    if user_id != current_user["id"] and not is_admin_user(current_user):
        raise HTTPException(status_code=403, detail="Access denied")
    doc = {
        "id": generate_uuid(),
        "name": data.get("name", "Document"), "file_name": data.get("file_name", ""),
        "file_data": data.get("file_data", ""), "document_type": data.get("document_type", "general"),
        "status": "processing", "notes": data.get("notes", ""),
        "uploaded_at": datetime.now(timezone.utc).isoformat(),
        "reviewed_at": None, "reviewed_by": None,
    }
    await db.users.update_one({"id": user_id}, {"$push": {"documents": doc}, "$set": {"updated_at": datetime.now(timezone.utc).isoformat()}})
    return doc


@router.put("/users/{user_id}/documents/{doc_id}")
async def update_user_document_status(user_id: str, doc_id: str, data: dict, current_user: dict = Depends(get_current_user)):
    allowed_status = {"processing", "approved", "rejected"}
    new_status = data.get("status")
    if new_status and new_status not in allowed_status:
        raise HTTPException(status_code=400, detail=f"Status must be one of: {', '.join(allowed_status)}")
    if "admin" not in current_user.get("roles", []) and "manager" not in current_user.get("roles", []):
        raise HTTPException(status_code=403, detail="Only admins/managers can update document status")
    update_fields = {}
    if new_status:
        update_fields["documents.$.status"] = new_status
    if data.get("notes"):
        update_fields["documents.$.notes"] = data["notes"]
    update_fields["documents.$.reviewed_at"] = datetime.now(timezone.utc).isoformat()
    update_fields["documents.$.reviewed_by"] = current_user["id"]
    await db.users.update_one({"id": user_id, "documents.id": doc_id}, {"$set": update_fields})
    return {"message": "Document status updated"}


@router.delete("/users/{user_id}/documents/{doc_id}")
async def delete_user_document(user_id: str, doc_id: str, current_user: dict = Depends(get_current_user)):
    if user_id != current_user["id"] and not is_admin_user(current_user):
        raise HTTPException(status_code=403, detail="Access denied")
    await db.users.update_one({"id": user_id}, {"$pull": {"documents": {"id": doc_id}}})
    return {"message": "Document deleted"}


# --- Batch & Import ---

@router.post("/users/batch-action")
async def batch_user_action(body: dict, current_user: dict = Depends(get_current_user)):
    if not is_admin_user(current_user):
        raise HTTPException(status_code=403, detail="Only admins can perform batch actions")
    action = body.get("action")
    user_ids = body.get("user_ids", [])
    if not user_ids:
        raise HTTPException(status_code=400, detail="No users selected")
    if action not in ["activate", "deactivate", "delete"]:
        raise HTTPException(status_code=400, detail="Invalid action. Use: activate, deactivate, delete")
    user_ids = [uid for uid in user_ids if uid != current_user["id"]]
    if not user_ids:
        raise HTTPException(status_code=400, detail="Cannot perform batch actions on yourself")
    count = 0
    now = datetime.now(timezone.utc).isoformat()
    if action == "activate":
        result = await db.users.update_many({"id": {"$in": user_ids}}, {"$set": {"is_active": True, "updated_at": now}})
        count = result.modified_count
    elif action == "deactivate":
        result = await db.users.update_many({"id": {"$in": user_ids}}, {"$set": {"is_active": False, "updated_at": now}})
        count = result.modified_count
    elif action == "delete":
        result = await db.users.delete_many({"id": {"$in": user_ids}})
        count = result.deleted_count
    return {"message": f"{action.title()}d {count} user(s)", "count": count}


@router.post("/users/bulk-import")
async def bulk_import_users(users_data: list, current_user: dict = Depends(get_current_user)):
    if not is_admin_user(current_user):
        raise HTTPException(status_code=403, detail="Only admins can import users")
    imported = 0
    skipped = 0
    errors = []
    for user_data in users_data:
        try:
            email = user_data.get("email", "").strip().lower()
            if not email:
                skipped += 1
                continue
            existing = await db.users.find_one({"email": email})
            if existing:
                skipped += 1
                continue
            ofo_major = user_data.get("ofo_major_group", "")
            role = "employee"
            if "MANAGERS" in ofo_major.upper():
                role = "manager"
            elif "PROFESSIONALS" in ofo_major.upper():
                role = "professional"
            elif "TECHNICIANS" in ofo_major.upper():
                role = "technician"
            elif "CLERICAL" in ofo_major.upper():
                role = "clerical"
            user = {
                "id": generate_uuid(), "email": email,
                "full_name": user_data.get("full_name", ""), "surname": user_data.get("surname", ""),
                "password_hash": None, "roles": [role],
                "division": user_data.get("division", ""), "department": user_data.get("department", ""),
                "position": user_data.get("position", ""),
                "is_active": True, "is_verified": False, "requires_password_setup": True,
                "ofo_major_group": ofo_major, "ofo_sub_major_group": user_data.get("ofo_sub_major_group", ""),
                "ofo_occupation": user_data.get("ofo_occupation", ""), "ofo_code": user_data.get("ofo_code", ""),
                "personnel_number": user_data.get("personnel_number", ""),
                "id_number": user_data.get("id_number", ""),
                "race": user_data.get("race", ""), "gender": user_data.get("gender", ""),
                "age": user_data.get("age"), "start_date": user_data.get("start_date", ""),
                "years_of_service": user_data.get("years_of_service"), "level": user_data.get("level", ""),
                "created_at": datetime.now(timezone.utc).isoformat(),
                "updated_at": datetime.now(timezone.utc).isoformat(),
            }
            await db.users.insert_one({**user})
            imported += 1
        except Exception as e:
            errors.append({"email": user_data.get("email", "unknown"), "error": str(e)})
    return {"imported": imported, "skipped": skipped, "errors": errors, "total_processed": imported + skipped + len(errors)}
