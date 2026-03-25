"""Reports and Dashboard Stats Router"""
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from typing import Optional, List
from datetime import datetime, timezone
import io

from routers import db, get_current_user, generate_uuid, generate_excel, is_admin_user

router = APIRouter(prefix="/api", tags=["reports"])


@router.get("/reports/interactive-data")
async def get_interactive_report_data(
    division: Optional[str] = None,
    divisions: Optional[str] = None,
    subgroup: Optional[str] = None,
    departments: Optional[str] = None,
    status: Optional[str] = None,
    race: Optional[str] = None,
    races: Optional[str] = None,
    age_min: Optional[int] = None,
    age_max: Optional[int] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
):
    """Returns comprehensive data for interactive reporting with demographic filtering."""

    # ── Build user query ──
    user_query = {}
    # Multi-value division filter
    div_list = []
    if divisions:
        div_list = [d.strip() for d in divisions.split(",") if d.strip()]
    elif division:
        div_list = [division]
    if div_list:
        user_query["division"] = {"$in": div_list} if len(div_list) > 1 else div_list[0]

    # Multi-value department filter
    dept_list = []
    if departments:
        dept_list = [d.strip() for d in departments.split(",") if d.strip()]
    elif subgroup:
        dept_list = [subgroup]
    if dept_list:
        user_query["department"] = {"$in": dept_list} if len(dept_list) > 1 else dept_list[0]

    # Multi-value race filter
    race_list = []
    if races:
        race_list = [r.strip() for r in races.split(",") if r.strip()]
    elif race:
        race_list = [race]
    if race_list:
        user_query["race"] = {"$in": race_list} if len(race_list) > 1 else race_list[0]

    # Age range filter
    if age_min is not None or age_max is not None:
        age_q = {}
        if age_min is not None:
            age_q["$gte"] = age_min
        if age_max is not None:
            age_q["$lte"] = age_max
        user_query["age"] = age_q

    # ── Users data ──
    users = await db.users.find(user_query, {"_id": 0, "password_hash": 0}).to_list(10000)
    user_ids = [u["id"] for u in users]

    # Division breakdown
    div_counts = {}
    dept_counts = {}
    role_counts = {}
    race_counts = {}
    age_ranges = {"18-25": 0, "26-35": 0, "36-45": 0, "46-55": 0, "56-65": 0, "65+": 0, "Unknown": 0}
    active_count = 0
    inactive_count = 0
    for u in users:
        d = u.get("division") or "Unassigned"
        div_counts[d] = div_counts.get(d, 0) + 1
        dept = u.get("department") or "Unassigned"
        dept_counts[dept] = dept_counts.get(dept, 0) + 1
        for r in u.get("roles", []):
            role_counts[r] = role_counts.get(r, 0) + 1
        rc = u.get("race") or "Unspecified"
        race_counts[rc] = race_counts.get(rc, 0) + 1
        age = u.get("age")
        if age is not None and isinstance(age, (int, float)):
            age = int(age)
            if age <= 25:
                age_ranges["18-25"] += 1
            elif age <= 35:
                age_ranges["26-35"] += 1
            elif age <= 45:
                age_ranges["36-45"] += 1
            elif age <= 55:
                age_ranges["46-55"] += 1
            elif age <= 65:
                age_ranges["56-65"] += 1
            else:
                age_ranges["65+"] += 1
        else:
            age_ranges["Unknown"] += 1
        if u.get("is_active"):
            active_count += 1
        else:
            inactive_count += 1

    # ── Build app query ──
    app_query = {}
    has_user_filter = bool(div_list or dept_list or race_list or age_min is not None or age_max is not None)
    if user_ids and has_user_filter:
        app_query["user_id"] = {"$in": user_ids}
    if status:
        app_query["status"] = status
    if date_from:
        app_query.setdefault("created_at", {})["$gte"] = date_from
    if date_to:
        app_query.setdefault("created_at", {})["$lte"] = date_to + "T23:59:59"

    # ── Bursary Applications ──
    bursary_apps = await db.applications.find(app_query, {"_id": 0}).to_list(10000)
    bursary_status = {}
    bursary_by_div = {}
    bursary_total_amount = 0
    for a in bursary_apps:
        s = a.get("status", "unknown")
        bursary_status[s] = bursary_status.get(s, 0) + 1
        # Match user's division
        applicant = next((u for u in users if u["id"] == a.get("user_id")), None)
        d = applicant.get("division", "Unknown") if applicant else "Unknown"
        bursary_by_div[d] = bursary_by_div.get(d, 0) + 1
        fi = a.get("financial_info", {})
        amt = fi.get("total_amount") or fi.get("amount_requested") or fi.get("bursary_amount") or 0
        bursary_total_amount += float(amt) if amt else 0

    # ── Training Applications ──
    training_apps = await db.training_applications.find(app_query, {"_id": 0}).to_list(10000)
    training_status = {}
    training_by_div = {}
    training_total_amount = 0
    training_by_provider = {}
    for a in training_apps:
        s = a.get("status", "unknown")
        training_status[s] = training_status.get(s, 0) + 1
        applicant = next((u for u in users if u["id"] == a.get("user_id")), None)
        d = applicant.get("division", "Unknown") if applicant else "Unknown"
        training_by_div[d] = training_by_div.get(d, 0) + 1
        ti = a.get("training_info", {})
        amt = ti.get("total_amount") or ti.get("amount_requested") or 0
        training_total_amount += float(amt) if amt else 0
        provider = ti.get("service_provider") or "Unknown"
        training_by_provider[provider] = training_by_provider.get(provider, 0) + 1

    # ── Expenses from applications ──
    expense_types = {"Flights": 0, "Accommodation": 0, "Car Hire/Shuttle": 0, "Catering": 0}
    total_app_expenses = 0
    expense_by_div = {}
    applicant_expenses = []

    for coll_name, apps in [("bursary", bursary_apps), ("training", training_apps)]:
        for a in apps:
            exp = a.get("additional_expenses", {})
            if not exp:
                continue
            flights = float(exp.get("flights", 0) or 0)
            accommodation = float(exp.get("accommodation", 0) or 0)
            car = float(exp.get("car_hire_or_shuttle", 0) or 0)
            catering = float(exp.get("catering", 0) or 0)
            total = flights + accommodation + car + catering
            if total > 0:
                expense_types["Flights"] += flights
                expense_types["Accommodation"] += accommodation
                expense_types["Car Hire/Shuttle"] += car
                expense_types["Catering"] += catering
                total_app_expenses += total
                applicant = next((u for u in users if u["id"] == a.get("user_id")), None)
                d = applicant.get("division", "Unknown") if applicant else "Unknown"
                expense_by_div[d] = expense_by_div.get(d, 0) + total
                applicant_expenses.append({
                    "applicant": a.get("applicant_name") or (applicant.get("full_name") if applicant else "Unknown"),
                    "type": coll_name,
                    "division": d,
                    "department": applicant.get("department", "Unknown") if applicant else "Unknown",
                    "flights": flights, "accommodation": accommodation,
                    "car_hire": car, "catering": catering,
                    "total": total,
                })

    # ── Standalone expenses ──
    standalone_query = {}
    if user_ids and has_user_filter:
        standalone_query["submitted_by"] = {"$in": user_ids}
    if date_from:
        standalone_query.setdefault("date", {})["$gte"] = date_from
    if date_to:
        standalone_query.setdefault("date", {})["$lte"] = date_to + "T23:59:59"
    standalone_expenses = await db.expenses.find(standalone_query, {"_id": 0}).to_list(10000)
    standalone_total = sum(float(e.get("amount", 0)) for e in standalone_expenses)
    standalone_by_cat = {}
    for e in standalone_expenses:
        cat = e.get("category", "other")
        standalone_by_cat[cat] = standalone_by_cat.get(cat, 0) + float(e.get("amount", 0))

    # ── Tickets ──
    ticket_query = {}
    if user_ids and has_user_filter:
        ticket_query["created_by"] = {"$in": user_ids}
    if date_from:
        ticket_query.setdefault("created_at", {})["$gte"] = date_from
    if date_to:
        ticket_query.setdefault("created_at", {})["$lte"] = date_to + "T23:59:59"
    tickets = await db.tickets.find(ticket_query, {"_id": 0}).to_list(10000)
    ticket_status = {}
    ticket_category = {}
    ticket_priority = {}
    for t in tickets:
        s = t.get("status", "unknown")
        ticket_status[s] = ticket_status.get(s, 0) + 1
        c = t.get("category", "other")
        ticket_category[c] = ticket_category.get(c, 0) + 1
        p = t.get("priority", "medium")
        ticket_priority[p] = ticket_priority.get(p, 0) + 1

    # ── Division, department, race lists for filters ──
    all_divisions = await db.users.distinct("division")
    all_divisions = [d for d in all_divisions if d]
    if div_list:
        all_departments = await db.users.distinct("department", {"division": {"$in": div_list}})
    else:
        all_departments = await db.users.distinct("department")
    all_departments = [d for d in all_departments if d]
    all_races = await db.users.distinct("race")
    all_races = [r for r in all_races if r]

    return {
        "filters": {
            "divisions": sorted(all_divisions),
            "departments": sorted(all_departments),
            "races": sorted(all_races),
        },
        "users": {
            "total": len(users),
            "active": active_count,
            "inactive": inactive_count,
            "by_division": [{"name": k, "value": v} for k, v in sorted(div_counts.items(), key=lambda x: -x[1])],
            "by_department": [{"name": k, "value": v} for k, v in sorted(dept_counts.items(), key=lambda x: -x[1])],
            "by_role": [{"name": k, "value": v} for k, v in sorted(role_counts.items(), key=lambda x: -x[1])],
            "by_race": [{"name": k, "value": v} for k, v in sorted(race_counts.items(), key=lambda x: -x[1])],
            "by_age_range": [{"name": k, "value": v} for k, v in age_ranges.items() if v > 0],
        },
        "bursary_applications": {
            "total": len(bursary_apps),
            "by_status": [{"name": k, "value": v} for k, v in bursary_status.items()],
            "by_division": [{"name": k, "value": v} for k, v in bursary_by_div.items()],
            "total_amount": bursary_total_amount,
        },
        "training_applications": {
            "total": len(training_apps),
            "by_status": [{"name": k, "value": v} for k, v in training_status.items()],
            "by_division": [{"name": k, "value": v} for k, v in training_by_div.items()],
            "total_amount": training_total_amount,
            "by_provider": [{"name": k, "value": v} for k, v in training_by_provider.items()],
        },
        "expenses": {
            "by_type": [{"name": k, "value": v} for k, v in expense_types.items() if v > 0],
            "by_division": [{"name": k, "value": v} for k, v in expense_by_div.items()],
            "by_applicant": applicant_expenses,
            "total_application_expenses": total_app_expenses,
            "standalone_total": standalone_total,
            "standalone_by_category": [{"name": k, "value": v} for k, v in standalone_by_cat.items()],
        },
        "tickets": {
            "total": len(tickets),
            "by_status": [{"name": k, "value": v} for k, v in ticket_status.items()],
            "by_category": [{"name": k, "value": v} for k, v in ticket_category.items()],
            "by_priority": [{"name": k, "value": v} for k, v in ticket_priority.items()],
        },
    }


@router.post("/reports/export-chart")
async def export_chart_data(
    body: dict,
    current_user: dict = Depends(get_current_user),
):
    """Export a single chart's data as XLSX"""
    chart_title = body.get("title", "Chart Data")
    chart_data = body.get("data", [])

    if not chart_data:
        raise HTTPException(status_code=400, detail="No data to export")

    excel = generate_excel(chart_data, title=chart_title[:31])
    fname = chart_title.replace(" ", "_").lower() + ".xlsx"

    return StreamingResponse(
        iter([excel.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )



@router.get("/reports/export-filtered")
async def export_filtered_data(
    data_type: str = "all",
    division: Optional[str] = None,
    divisions: Optional[str] = None,
    subgroup: Optional[str] = None,
    departments: Optional[str] = None,
    status: Optional[str] = None,
    race: Optional[str] = None,
    races: Optional[str] = None,
    age_min: Optional[int] = None,
    age_max: Optional[int] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
):
    """Export filtered data as XLSX"""
    user_query = {}
    div_list = [d.strip() for d in (divisions or "").split(",") if d.strip()] if divisions else ([division] if division else [])
    dept_list = [d.strip() for d in (departments or "").split(",") if d.strip()] if departments else ([subgroup] if subgroup else [])
    race_list = [r.strip() for r in (races or "").split(",") if r.strip()] if races else ([race] if race else [])
    if div_list:
        user_query["division"] = {"$in": div_list} if len(div_list) > 1 else div_list[0]
    if dept_list:
        user_query["department"] = {"$in": dept_list} if len(dept_list) > 1 else dept_list[0]
    if race_list:
        user_query["race"] = {"$in": race_list} if len(race_list) > 1 else race_list[0]
    if age_min is not None or age_max is not None:
        age_q = {}
        if age_min is not None:
            age_q["$gte"] = age_min
        if age_max is not None:
            age_q["$lte"] = age_max
        user_query["age"] = age_q

    has_user_filter = bool(div_list or dept_list or race_list or age_min is not None or age_max is not None)

    users = await db.users.find(user_query, {"_id": 0, "password_hash": 0}).to_list(10000)
    user_ids = [u["id"] for u in users]
    user_map = {u["id"]: u for u in users}

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    wb.remove(wb.active)

    def write_sheet(ws, title, rows):
        if not rows:
            ws.append(["No data"])
            return
        headers = list(rows[0].keys())
        header_fill = PatternFill(start_color="0056B3", end_color="0056B3", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h.replace("_", " ").title())
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        alt = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
        for ri, row in enumerate(rows, 2):
            for ci, h in enumerate(headers, 1):
                val = row.get(h, "")
                if isinstance(val, (list, dict)):
                    val = str(val)
                cell = ws.cell(row=ri, column=ci, value=val)
                if ri % 2 == 0:
                    cell.fill = alt
        for ci, h in enumerate(headers, 1):
            mx = max(len(str(h)), max((len(str(r.get(h, ""))) for r in rows), default=0))
            ws.column_dimensions[get_column_letter(ci)].width = min(mx + 4, 50)

    if data_type in ("all", "users"):
        ws = wb.create_sheet("Users")
        rows = [{"Name": u.get("full_name", ""), "Email": u.get("email", ""),
                 "Division": u.get("division", ""), "Department": u.get("department", ""),
                 "Race": u.get("race", ""), "Age": u.get("age", ""),
                 "Roles": ", ".join(u.get("roles", [])), "Active": "Yes" if u.get("is_active") else "No"}
                for u in users]
        write_sheet(ws, "Users", rows)

    if data_type in ("all", "applications"):
        app_q = {}
        if user_ids and has_user_filter:
            app_q["user_id"] = {"$in": user_ids}
        if status:
            app_q["status"] = status
        if date_from:
            app_q.setdefault("created_at", {})["$gte"] = date_from
        if date_to:
            app_q.setdefault("created_at", {})["$lte"] = date_to + "T23:59:59"

        bursary = await db.applications.find(app_q, {"_id": 0}).to_list(10000)
        ws = wb.create_sheet("Bursary Applications")
        rows = []
        for a in bursary:
            u = user_map.get(a.get("user_id"), {})
            fi = a.get("financial_info", {})
            exp = a.get("additional_expenses", {})
            rows.append({
                "Applicant": a.get("applicant_name") or u.get("full_name", ""),
                "Email": a.get("user_email") or u.get("email", ""),
                "Division": u.get("division", ""),
                "Department": u.get("department", ""),
                "Status": a.get("status", ""),
                "Requested Amount": fi.get("total_amount") or fi.get("amount_requested") or 0,
                "Flights": exp.get("flights", 0),
                "Accommodation": exp.get("accommodation", 0),
                "Car Hire": exp.get("car_hire_or_shuttle", 0),
                "Catering": exp.get("catering", 0),
                "Created": str(a.get("created_at", ""))[:10],
            })
        write_sheet(ws, "Bursary Applications", rows)

        training = await db.training_applications.find(app_q, {"_id": 0}).to_list(10000)
        ws = wb.create_sheet("Training Applications")
        rows = []
        for a in training:
            u = user_map.get(a.get("user_id"), {})
            ti = a.get("training_info", {})
            exp = a.get("additional_expenses", {})
            rows.append({
                "Applicant": a.get("applicant_name") or u.get("full_name", ""),
                "Division": u.get("division", ""),
                "Department": u.get("department", ""),
                "Status": a.get("status", ""),
                "Provider": ti.get("service_provider", ""),
                "Amount": ti.get("total_amount") or ti.get("amount_requested") or 0,
                "Flights": exp.get("flights", 0),
                "Accommodation": exp.get("accommodation", 0),
                "Car Hire": exp.get("car_hire_or_shuttle", 0),
                "Catering": exp.get("catering", 0),
                "Created": str(a.get("created_at", ""))[:10],
            })
        write_sheet(ws, "Training Applications", rows)

    if data_type in ("all", "expenses"):
        exp_q = {}
        if user_ids and has_user_filter:
            exp_q["submitted_by"] = {"$in": user_ids}
        if date_from:
            exp_q.setdefault("date", {})["$gte"] = date_from
        if date_to:
            exp_q.setdefault("date", {})["$lte"] = date_to + "T23:59:59"
        expenses = await db.expenses.find(exp_q, {"_id": 0}).to_list(10000)
        ws = wb.create_sheet("Expenses")
        rows = [{"Title": e.get("title", ""), "Category": e.get("category", ""),
                 "Amount": e.get("amount", 0), "Status": e.get("status", ""),
                 "Vendor": e.get("vendor", ""), "Date": str(e.get("date", ""))[:10]}
                for e in expenses]
        write_sheet(ws, "Expenses", rows)

    if data_type in ("all", "tickets"):
        tkt_q = {}
        if user_ids and has_user_filter:
            tkt_q["created_by"] = {"$in": user_ids}
        if date_from:
            tkt_q.setdefault("created_at", {})["$gte"] = date_from
        if date_to:
            tkt_q.setdefault("created_at", {})["$lte"] = date_to + "T23:59:59"
        tkts = await db.tickets.find(tkt_q, {"_id": 0}).to_list(10000)
        ws = wb.create_sheet("Tickets")
        rows = [{"Title": t.get("title", ""), "Category": t.get("category", ""),
                 "Priority": t.get("priority", ""), "Status": t.get("status", ""),
                 "Created By": t.get("created_by", ""), "Created": str(t.get("created_at", ""))[:10]}
                for t in tkts]
        write_sheet(ws, "Tickets", rows)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    fname = f"sentech_report_{data_type}"
    if division:
        fname += f"_{division.replace(' ', '_')}"
    fname += ".xlsx"

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


@router.get("/dashboard/stats")
async def get_dashboard_stats(current_user: dict = Depends(get_current_user)):
    is_admin = "admin" in current_user.get("roles", []) or "super_admin" in current_user.get("roles", [])
    if is_admin:
        total_applications = await db.applications.count_documents({})
        pending_applications = await db.applications.count_documents({"status": "pending"})
        approved_applications = await db.applications.count_documents({"status": "approved"})
        training_applications = await db.training_applications.count_documents({})
        open_tickets = await db.tickets.count_documents({"status": "open"})
        active_users = await db.users.count_documents({"is_active": True})
    else:
        total_applications = await db.applications.count_documents({"user_id": current_user["id"]})
        pending_applications = await db.applications.count_documents({"user_id": current_user["id"], "status": "pending"})
        approved_applications = await db.applications.count_documents({"user_id": current_user["id"], "status": "approved"})
        training_applications = await db.training_applications.count_documents({"user_id": current_user["id"]})
        open_tickets = await db.tickets.count_documents({"created_by": current_user["id"], "status": "open"})
        active_users = 0

    unread_notifications = await db.notifications.count_documents({"user_id": current_user["id"], "is_read": False})

    return {
        "total_applications": total_applications,
        "pending_applications": pending_applications,
        "approved_applications": approved_applications,
        "training_applications": training_applications,
        "open_tickets": open_tickets,
        "unread_notifications": unread_notifications,
        "active_users": active_users,
    }


@router.get("/dashboard/recent-activity")
async def get_recent_activity(current_user: dict = Depends(get_current_user)):
    is_admin = "admin" in current_user.get("roles", []) or "super_admin" in current_user.get("roles", [])
    activities = []

    notifications = await db.notifications.find(
        {"user_id": current_user["id"]}, {"_id": 0}
    ).sort("created_at", -1).limit(10).to_list(10)
    for n in notifications:
        activities.append({
            "id": n.get("id", ""), "type": n.get("type", "notification"),
            "title": n.get("title", "Notification"),
            "description": n.get("message", ""),
            "created_at": n.get("created_at", ""), "icon": "bell",
        })

    app_query = {} if is_admin else {"user_id": current_user["id"]}
    recent_apps = await db.applications.find(app_query, {"_id": 0}).sort("created_at", -1).limit(5).to_list(5)
    for a in recent_apps:
        applicant = a.get("applicant_name") or a.get("full_name") or "Unknown"
        activities.append({
            "id": a.get("id", ""), "type": "application",
            "title": f"Bursary Application - {a.get('status', 'submitted').title()}",
            "description": f"{applicant} - {a.get('institution', 'N/A')}",
            "created_at": a.get("created_at", ""), "icon": "file-text",
        })

    ta_query = {} if is_admin else {"user_id": current_user["id"]}
    recent_tas = await db.training_applications.find(ta_query, {"_id": 0}).sort("created_at", -1).limit(5).to_list(5)
    for t in recent_tas:
        applicant = t.get("applicant_name") or t.get("full_name") or "Unknown"
        activities.append({
            "id": t.get("id", ""), "type": "training_application",
            "title": f"Training Application - {t.get('status', 'submitted').title()}",
            "description": f"{applicant} - {t.get('training_title', 'N/A')}",
            "created_at": t.get("created_at", ""), "icon": "graduation-cap",
        })

    tkt_query = {} if is_admin else {"created_by": current_user["id"]}
    recent_tickets = await db.tickets.find(tkt_query, {"_id": 0}).sort("created_at", -1).limit(5).to_list(5)
    for tk in recent_tickets:
        activities.append({
            "id": tk.get("id", ""), "type": "ticket",
            "title": f"Ticket: {tk.get('title', 'Untitled')}",
            "description": f"Status: {tk.get('status', 'open').title()} - Priority: {tk.get('priority', 'medium').title()}",
            "created_at": tk.get("created_at", ""), "icon": "ticket",
        })

    activities.sort(key=lambda x: x.get("created_at", ""), reverse=True)
    return activities[:15]


@router.get("/dashboard/report-summary")
async def get_dashboard_report_summary(current_user: dict = Depends(get_current_user)):
    total_users = await db.users.count_documents({})
    active_users = await db.users.count_documents({"is_active": True})
    total_bursary = await db.applications.count_documents({})
    approved_bursary = await db.applications.count_documents({"status": "approved"})
    total_training = await db.training_applications.count_documents({})
    approved_training = await db.training_applications.count_documents({"status": "approved"})
    total_expenses = 0
    expenses_cursor = db.expenses.find({}, {"_id": 0, "amount": 1})
    async for exp in expenses_cursor:
        total_expenses += exp.get("amount", 0)
    open_tickets = await db.tickets.count_documents({"status": "open"})
    closed_tickets = await db.tickets.count_documents({"status": "closed"})

    return {
        "total_users": total_users, "active_users": active_users,
        "total_bursary_applications": total_bursary, "approved_bursary_applications": approved_bursary,
        "total_training_applications": total_training, "approved_training_applications": approved_training,
        "total_expenses": total_expenses,
        "open_tickets": open_tickets, "closed_tickets": closed_tickets,
    }
