from fastapi import FastAPI, APIRouter, Depends, HTTPException, status, Request, UploadFile, File, Form
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import RedirectResponse, StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import asyncio
import os
import logging
import sys
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
from passlib.context import CryptContext
from jose import JWTError, jwt
import secrets
import msal
import requests
import io
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# Add models to path
sys.path.insert(0, str(ROOT_DIR))

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

app = FastAPI(title="Sentech Bursary Management System")
api_router = APIRouter(prefix="/api")

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
security = HTTPBearer()

SECRET_KEY = os.getenv("JWT_SECRET_KEY", "your-secret-key-change-in-production-min-32-chars")
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 1440

MICROSOFT_CLIENT_ID = os.getenv("MICROSOFT_CLIENT_ID")
MICROSOFT_TENANT_ID = os.getenv("MICROSOFT_TENANT_ID")
MICROSOFT_CLIENT_SECRET = os.getenv("MICROSOFT_CLIENT_SECRET")
MICROSOFT_AUTHORITY = f"https://login.microsoftonline.com/{MICROSOFT_TENANT_ID}"
MICROSOFT_SCOPE = ["User.Read"]

FRONTEND_URL = os.getenv("REACT_APP_BACKEND_URL", "http://localhost:3000").replace("/api", "")

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Helper functions
def generate_uuid():
    return str(uuid.uuid4())

def current_time():
    return datetime.now(timezone.utc)

def serialize_datetime(obj):
    if isinstance(obj, datetime):
        return obj.isoformat()
    return obj

# ============== MODELS ==============

class User(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=generate_uuid)
    email: EmailStr
    full_name: str
    surname: Optional[str] = None
    password_hash: Optional[str] = None
    roles: List[str] = ["student"]
    team_id: Optional[str] = None
    department: Optional[str] = None
    division: Optional[str] = None
    position: Optional[str] = None
    phone: Optional[str] = None
    avatar_url: Optional[str] = None
    is_active: bool = True
    is_verified: bool = False
    requires_password_setup: bool = False  # For first-time login
    student_id: Optional[str] = None  # deprecated
    permissions: Dict[str, Any] = {}
    bio: Optional[str] = None
    surname: Optional[str] = None
    # OFO Fields
    ofo_major_group: Optional[str] = None
    ofo_sub_major_group: Optional[str] = None
    ofo_occupation: Optional[str] = None
    ofo_code: Optional[str] = None
    # Personal Details
    personnel_number: Optional[str] = None
    id_number: Optional[str] = None
    race: Optional[str] = None
    gender: Optional[str] = None
    age: Optional[int] = None
    # Employment Details
    start_date: Optional[str] = None
    years_of_service: Optional[float] = None
    level: Optional[str] = None
    created_at: datetime = Field(default_factory=current_time)
    updated_at: datetime = Field(default_factory=current_time)

class UserCreate(BaseModel):
    email: EmailStr
    full_name: str
    password: Optional[str] = None
    surname: Optional[str] = None
    student_id: Optional[str] = None
    roles: List[str] = ["student"]
    team_id: Optional[str] = None
    department: Optional[str] = None
    division: Optional[str] = None
    position: Optional[str] = None
    phone: Optional[str] = None
    requires_password_setup: bool = False
    # OFO Fields
    ofo_major_group: Optional[str] = None
    ofo_sub_major_group: Optional[str] = None
    ofo_occupation: Optional[str] = None
    ofo_code: Optional[str] = None
    # Personal Details
    personnel_number: Optional[str] = None
    id_number: Optional[str] = None
    race: Optional[str] = None
    gender: Optional[str] = None
    age: Optional[int] = None
    # Employment Details
    start_date: Optional[str] = None
    years_of_service: Optional[float] = None
    level: Optional[str] = None

class UserUpdate(BaseModel):
    full_name: Optional[str] = None
    surname: Optional[str] = None
    bio: Optional[str] = None
    roles: Optional[List[str]] = None
    team_id: Optional[str] = None
    department: Optional[str] = None
    division: Optional[str] = None
    position: Optional[str] = None
    phone: Optional[str] = None
    avatar_url: Optional[str] = None
    is_active: Optional[bool] = None
    requires_password_setup: Optional[bool] = None
    # OFO Fields
    ofo_major_group: Optional[str] = None
    ofo_sub_major_group: Optional[str] = None
    ofo_occupation: Optional[str] = None
    ofo_code: Optional[str] = None
    # Personal Details
    personnel_number: Optional[str] = None
    id_number: Optional[str] = None
    race: Optional[str] = None
    gender: Optional[str] = None
    age: Optional[int] = None
    email: Optional[str] = None
    # Employment Details
    start_date: Optional[str] = None
    years_of_service: Optional[float] = None
    level: Optional[str] = None

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class PasswordSetup(BaseModel):
    email: EmailStr
    new_password: str

class Division(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=generate_uuid)
    name: str
    description: Optional[str] = None
    is_active: bool = True
    created_at: datetime = Field(default_factory=current_time)
    updated_at: datetime = Field(default_factory=current_time)

class Department(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=generate_uuid)
    name: str
    division_id: str
    description: Optional[str] = None
    is_active: bool = True
    created_at: datetime = Field(default_factory=current_time)
    updated_at: datetime = Field(default_factory=current_time)

class TokenResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"
    user: Dict[str, Any]

class Team(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=generate_uuid)
    name: str
    description: Optional[str] = None
    leader_id: Optional[str] = None
    member_ids: List[str] = []
    department: Optional[str] = None
    color: Optional[str] = "#0056B3"
    is_active: bool = True
    created_at: datetime = Field(default_factory=current_time)
    updated_at: datetime = Field(default_factory=current_time)

class TeamCreate(BaseModel):
    name: str
    description: Optional[str] = None
    leader_id: Optional[str] = None
    member_ids: List[str] = []
    department: Optional[str] = None
    color: Optional[str] = "#0056B3"

class TeamUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    leader_id: Optional[str] = None
    member_ids: Optional[List[str]] = None
    department: Optional[str] = None
    color: Optional[str] = None
    is_active: Optional[bool] = None

class Meeting(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=generate_uuid)
    title: str
    description: Optional[str] = None
    meeting_type: str = "general"
    meeting_link: Optional[str] = None
    meeting_id: Optional[str] = None
    meeting_password: Optional[str] = None
    start_time: datetime
    end_time: datetime
    timezone: str = "Africa/Johannesburg"
    organizer_id: str
    attendee_ids: List[str] = []
    status: str = "scheduled"
    notes: Optional[str] = None
    project_id: Optional[str] = None
    created_at: datetime = Field(default_factory=current_time)
    updated_at: datetime = Field(default_factory=current_time)

class MeetingCreate(BaseModel):
    title: str
    description: Optional[str] = None
    meeting_type: str = "general"
    meeting_link: Optional[str] = None
    meeting_id: Optional[str] = None
    meeting_password: Optional[str] = None
    start_time: datetime
    end_time: datetime
    timezone: str = "Africa/Johannesburg"
    attendee_ids: List[str] = []
    project_id: Optional[str] = None

class MeetingUpdate(BaseModel):
    title: Optional[str] = None
    description: Optional[str] = None
    meeting_type: Optional[str] = None
    meeting_link: Optional[str] = None
    start_time: Optional[datetime] = None
    end_time: Optional[datetime] = None
    status: Optional[str] = None
    notes: Optional[str] = None
    attendee_ids: Optional[List[str]] = None

class Note(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=generate_uuid)
    title: str
    content: str
    created_by: str
    folder_id: Optional[str] = None
    tags: List[str] = []
    color: Optional[str] = None
    is_pinned: bool = False
    is_shared: bool = False
    shared_with: List[str] = []
    shared_with_teams: List[str] = []
    created_at: datetime = Field(default_factory=current_time)
    updated_at: datetime = Field(default_factory=current_time)

class NoteCreate(BaseModel):
    title: str
    content: str
    folder_id: Optional[str] = None
    tags: List[str] = []
    color: Optional[str] = None
    is_pinned: bool = False
    is_shared: bool = False
    shared_with: List[str] = []
    shared_with_teams: List[str] = []

class NoteUpdate(BaseModel):
    title: Optional[str] = None
    content: Optional[str] = None
    folder_id: Optional[str] = None
    tags: Optional[List[str]] = None
    color: Optional[str] = None
    is_pinned: Optional[bool] = None
    is_shared: Optional[bool] = None
    shared_with: Optional[List[str]] = None
    shared_with_teams: Optional[List[str]] = None

class NoteFolder(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=generate_uuid)
    name: str
    created_by: str
    parent_id: Optional[str] = None
    color: Optional[str] = None
    is_shared: bool = False
    shared_with: List[str] = []
    shared_with_teams: List[str] = []
    created_at: datetime = Field(default_factory=current_time)
    updated_at: datetime = Field(default_factory=current_time)

class NoteFolderCreate(BaseModel):
    name: str
    parent_id: Optional[str] = None
    color: Optional[str] = None
    is_shared: bool = False
    shared_with: List[str] = []
    shared_with_teams: List[str] = []

class Message(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=generate_uuid)
    conversation_id: str
    sender_id: str
    content: str
    message_type: str = "text"
    attachments: List[Dict[str, str]] = []
    is_read: bool = False
    read_by: List[str] = []
    is_deleted: bool = False
    created_at: datetime = Field(default_factory=current_time)
    updated_at: datetime = Field(default_factory=current_time)

class MessageCreate(BaseModel):
    conversation_id: Optional[str] = None
    content: str
    message_type: str = "text"
    attachments: List[Dict[str, str]] = []

class Conversation(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=generate_uuid)
    name: Optional[str] = None
    type: str = "direct"
    participant_ids: List[str]
    created_by: str
    last_message_id: Optional[str] = None
    last_message_at: Optional[datetime] = None
    is_archived: bool = False
    unread_count: Dict[str, int] = {}
    created_at: datetime = Field(default_factory=current_time)
    updated_at: datetime = Field(default_factory=current_time)

class ConversationCreate(BaseModel):
    name: Optional[str] = None
    type: str = "direct"
    participant_ids: List[str]

class Expense(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=generate_uuid)
    title: str
    description: Optional[str] = None
    amount: float
    currency: str = "ZAR"
    category: str
    date: datetime
    submitted_by: str
    user_id: Optional[str] = None
    project_id: Optional[str] = None
    sponsor_id: Optional[str] = None
    status: str = "pending"
    approved_by: Optional[str] = None
    approved_at: Optional[datetime] = None
    receipt_url: Optional[str] = None
    notes: Optional[str] = None
    vendor: Optional[str] = None
    created_at: datetime = Field(default_factory=current_time)
    updated_at: datetime = Field(default_factory=current_time)

class ExpenseCreate(BaseModel):
    title: str
    description: Optional[str] = None
    amount: float
    currency: str = "ZAR"
    category: str
    date: datetime
    user_id: Optional[str] = None
    project_id: Optional[str] = None
    sponsor_id: Optional[str] = None
    receipt_url: Optional[str] = None
    notes: Optional[str] = None
    vendor: Optional[str] = None

class ExpenseUpdate(BaseModel):
    title: Optional[str] = None
    description: Optional[str] = None
    amount: Optional[float] = None
    category: Optional[str] = None
    date: Optional[datetime] = None
    status: Optional[str] = None
    notes: Optional[str] = None

class Ticket(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=generate_uuid)
    title: str
    description: str
    category: str = "general"
    priority: str = "medium"
    status: str = "open"
    created_by: str
    assigned_to: Optional[str] = None
    team_id: Optional[str] = None
    project_id: Optional[str] = None
    tags: List[str] = []
    attachments: List[Dict[str, str]] = []
    resolution: Optional[str] = None
    resolved_at: Optional[datetime] = None
    closed_at: Optional[datetime] = None
    created_at: datetime = Field(default_factory=current_time)
    updated_at: datetime = Field(default_factory=current_time)

class TicketCreate(BaseModel):
    title: str
    description: str
    category: str = "general"
    priority: str = "medium"
    project_id: Optional[str] = None
    tags: List[str] = []

class TicketUpdate(BaseModel):
    title: Optional[str] = None
    description: Optional[str] = None
    category: Optional[str] = None
    priority: Optional[str] = None
    status: Optional[str] = None
    assigned_to: Optional[str] = None
    resolution: Optional[str] = None

class TicketComment(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=generate_uuid)
    ticket_id: str
    user_id: str
    content: str
    is_internal: bool = False
    attachments: List[Dict[str, str]] = []
    created_at: datetime = Field(default_factory=current_time)
    updated_at: datetime = Field(default_factory=current_time)

class TicketCommentCreate(BaseModel):
    ticket_id: str
    content: str
    is_internal: bool = False

class FileModel(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=generate_uuid)
    name: str
    original_name: str
    file_type: str
    mime_type: str
    size: int
    url: str
    folder_id: Optional[str] = None
    uploaded_by: str
    project_id: Optional[str] = None
    tags: List[str] = []
    description: Optional[str] = None
    is_shared: bool = False
    shared_with: List[str] = []
    shared_with_teams: List[str] = []
    created_at: datetime = Field(default_factory=current_time)
    updated_at: datetime = Field(default_factory=current_time)

class FileCreate(BaseModel):
    name: str
    original_name: str
    file_type: str
    mime_type: str
    size: int
    url: str
    folder_id: Optional[str] = None
    project_id: Optional[str] = None
    tags: List[str] = []
    description: Optional[str] = None

class FileUpdate(BaseModel):
    name: Optional[str] = None
    folder_id: Optional[str] = None
    tags: Optional[List[str]] = None
    description: Optional[str] = None
    is_shared: Optional[bool] = None
    shared_with: Optional[List[str]] = None
    shared_with_teams: Optional[List[str]] = None

class Folder(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=generate_uuid)
    name: str
    parent_id: Optional[str] = None
    created_by: str
    project_id: Optional[str] = None
    color: Optional[str] = None
    is_shared: bool = False
    shared_with: List[str] = []
    shared_with_teams: List[str] = []
    created_at: datetime = Field(default_factory=current_time)
    updated_at: datetime = Field(default_factory=current_time)

class FolderCreate(BaseModel):
    name: str
    parent_id: Optional[str] = None
    project_id: Optional[str] = None
    color: Optional[str] = None

class FolderUpdate(BaseModel):
    name: Optional[str] = None
    parent_id: Optional[str] = None
    color: Optional[str] = None
    is_shared: Optional[bool] = None
    shared_with: Optional[List[str]] = None

class Task(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=generate_uuid)
    title: str
    description: Optional[str] = None
    status: str = "todo"
    priority: str = "medium"
    assigned_to: Optional[str] = None
    assignee_name: Optional[str] = None
    due_date: Optional[datetime] = None
    start_date: Optional[datetime] = None
    completed_at: Optional[datetime] = None
    project_id: Optional[str] = None
    project_name: Optional[str] = None
    created_by: str
    tags: List[str] = []
    progress: int = 0
    attendance: List[Dict[str, Any]] = []  # Training module attendance records
    comments: List[Dict[str, Any]] = []  # Training module comments
    images: List[Dict[str, Any]] = []  # Training module images
    created_at: datetime = Field(default_factory=current_time)
    updated_at: datetime = Field(default_factory=current_time)

class TaskCreate(BaseModel):
    model_config = ConfigDict(extra="ignore")
    title: str
    description: Optional[str] = None
    status: str = "todo"
    priority: str = "medium"
    assigned_to: Optional[str] = None
    assignee_name: Optional[str] = None
    due_date: Optional[datetime] = None
    start_date: Optional[datetime] = None
    project_id: Optional[str] = None
    project_name: Optional[str] = None
    tags: List[str] = []
    attendance: List[Dict[str, Any]] = []
    comments: List[Dict[str, Any]] = []
    images: List[Dict[str, Any]] = []

class TaskUpdate(BaseModel):
    title: Optional[str] = None
    description: Optional[str] = None
    status: Optional[str] = None
    priority: Optional[str] = None
    assigned_to: Optional[str] = None
    due_date: Optional[datetime] = None
    start_date: Optional[datetime] = None
    project_id: Optional[str] = None
    tags: Optional[List[str]] = None
    progress: Optional[int] = None
    attendance: Optional[List[Dict[str, Any]]] = None
    comments: Optional[List[Dict[str, Any]]] = None
    images: Optional[List[Dict[str, Any]]] = None

class Project(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=generate_uuid)
    name: str
    description: Optional[str] = None
    status: str = "active"
    project_type: str = "internal"
    start_date: Optional[datetime] = None
    end_date: Optional[datetime] = None
    deadline: Optional[datetime] = None
    budget: float = 0.0
    spent: float = 0.0
    currency: str = "ZAR"
    has_budget: bool = True
    client_id: Optional[str] = None
    client_name: Optional[str] = None
    sponsor_id: Optional[str] = None
    team_id: Optional[str] = None
    manager_id: Optional[str] = None
    member_ids: List[str] = []
    tags: List[str] = []
    color: Optional[str] = "#0056B3"
    progress: int = 0
    priority: str = "medium"
    created_by: str
    created_at: datetime = Field(default_factory=current_time)
    updated_at: datetime = Field(default_factory=current_time)

class ProjectCreate(BaseModel):
    name: str
    description: Optional[str] = None
    status: str = "active"
    project_type: str = "internal"
    start_date: Optional[datetime] = None
    end_date: Optional[datetime] = None
    deadline: Optional[datetime] = None
    budget: float = 0.0
    has_budget: bool = True
    client_id: Optional[str] = None
    sponsor_id: Optional[str] = None
    manager_id: Optional[str] = None
    member_ids: List[str] = []
    tags: List[str] = []
    priority: str = "medium"

class ProjectUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    status: Optional[str] = None
    start_date: Optional[datetime] = None
    end_date: Optional[datetime] = None
    budget: Optional[float] = None
    spent: Optional[float] = None
    progress: Optional[int] = None
    priority: Optional[str] = None
    member_ids: Optional[List[str]] = None

class Client(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=generate_uuid)
    name: str
    company: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    address: Optional[str] = None
    contact_person: Optional[str] = None
    notes: Optional[str] = None
    is_active: bool = True
    created_by: str
    created_at: datetime = Field(default_factory=current_time)
    updated_at: datetime = Field(default_factory=current_time)

class ClientCreate(BaseModel):
    name: str
    company: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    address: Optional[str] = None
    contact_person: Optional[str] = None
    notes: Optional[str] = None

class Lead(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=generate_uuid)
    name: str
    email: Optional[str] = None
    phone: Optional[str] = None
    company: Optional[str] = None
    position: Optional[str] = None
    source: str = "website"
    status: str = "new"
    value: float = 0.0
    currency: str = "ZAR"
    assigned_to: Optional[str] = None
    assignee_name: Optional[str] = None
    notes: Optional[str] = None
    tags: List[str] = []
    last_contact_date: Optional[datetime] = None
    next_follow_up: Optional[datetime] = None
    created_by: str
    created_at: datetime = Field(default_factory=current_time)
    updated_at: datetime = Field(default_factory=current_time)

class LeadCreate(BaseModel):
    name: str
    email: Optional[str] = None
    phone: Optional[str] = None
    company: Optional[str] = None
    position: Optional[str] = None
    source: str = "website"
    status: str = "new"
    value: float = 0.0
    assigned_to: Optional[str] = None
    notes: Optional[str] = None
    tags: List[str] = []

class LeadUpdate(BaseModel):
    name: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    company: Optional[str] = None
    source: Optional[str] = None
    status: Optional[str] = None
    value: Optional[float] = None
    assigned_to: Optional[str] = None
    notes: Optional[str] = None
    tags: Optional[List[str]] = None

class Prospect(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=generate_uuid)
    name: str
    email: Optional[str] = None
    phone: Optional[str] = None
    company: Optional[str] = None
    position: Optional[str] = None
    source: str = "website"
    status: str = "identified"
    interest_level: str = "medium"
    budget_range: Optional[str] = None
    assigned_to: Optional[str] = None
    notes: Optional[str] = None
    tags: List[str] = []
    next_action: Optional[str] = None
    next_action_date: Optional[datetime] = None
    created_by: str
    created_at: datetime = Field(default_factory=current_time)
    updated_at: datetime = Field(default_factory=current_time)

class ProspectCreate(BaseModel):
    name: str
    email: Optional[str] = None
    phone: Optional[str] = None
    company: Optional[str] = None
    source: str = "website"
    interest_level: str = "medium"
    assigned_to: Optional[str] = None
    notes: Optional[str] = None

class ProspectUpdate(BaseModel):
    name: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    company: Optional[str] = None
    status: Optional[str] = None
    interest_level: Optional[str] = None
    assigned_to: Optional[str] = None
    notes: Optional[str] = None
    next_action: Optional[str] = None
    next_action_date: Optional[datetime] = None

class Sponsor(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=generate_uuid)
    name: str
    organization: str
    email: Optional[str] = None
    phone: Optional[str] = None
    address: Optional[str] = None
    website: Optional[str] = None
    contact_person: Optional[str] = None
    contact_email: Optional[str] = None
    contact_phone: Optional[str] = None
    total_contribution: float = 0.0
    currency: str = "ZAR"
    active_bursaries: int = 0
    status: str = "active"
    type: str = "corporate"
    bbbee_level: Optional[int] = None
    notes: Optional[str] = None
    tags: List[str] = []
    created_by: str
    created_at: datetime = Field(default_factory=current_time)
    updated_at: datetime = Field(default_factory=current_time)

class SponsorCreate(BaseModel):
    name: str
    organization: str
    email: Optional[str] = None
    phone: Optional[str] = None
    contact_person: Optional[str] = None
    contact_email: Optional[str] = None
    total_contribution: float = 0.0
    type: str = "corporate"
    bbbee_level: Optional[int] = None

class SponsorUpdate(BaseModel):
    name: Optional[str] = None
    organization: Optional[str] = None
    email: Optional[str] = None
    contact_person: Optional[str] = None
    total_contribution: Optional[float] = None
    active_bursaries: Optional[int] = None
    status: Optional[str] = None
    bbbee_level: Optional[int] = None

class SponsorContact(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=generate_uuid)
    sponsor_id: str
    name: str
    email: Optional[str] = None
    phone: Optional[str] = None
    position: Optional[str] = None
    is_primary: bool = False
    notes: Optional[str] = None
    created_by: str
    created_at: datetime = Field(default_factory=current_time)
    updated_at: datetime = Field(default_factory=current_time)

class SponsorContactCreate(BaseModel):
    sponsor_id: str
    name: str
    email: Optional[str] = None
    phone: Optional[str] = None
    position: Optional[str] = None
    is_primary: bool = False

class BursaryApplication(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=generate_uuid)
    user_id: str
    user_email: str
    status: str = "draft"
    current_step: int = 1
    personal_info: Dict[str, Any] = {}
    academic_info: Dict[str, Any] = {}
    financial_info: Dict[str, Any] = {}
    documents: Dict[str, Any] = {}
    created_at: datetime = Field(default_factory=current_time)
    updated_at: datetime = Field(default_factory=current_time)
    submitted_at: Optional[datetime] = None

class ApplicationCreate(BaseModel):
    status: Optional[str] = "draft"
    current_step: Optional[int] = 1
    personal_info: Optional[Dict[str, Any]] = {}
    academic_info: Optional[Dict[str, Any]] = {}
    academic_bursary_info: Optional[Dict[str, Any]] = {}
    employment_info: Optional[Dict[str, Any]] = {}
    financial_info: Optional[Dict[str, Any]] = {}
    documents: Optional[Dict[str, Any]] = {}

class ApplicationUpdate(BaseModel):
    current_step: Optional[int] = None
    status: Optional[str] = None
    status_note: Optional[str] = None
    status_updated_at: Optional[str] = None
    personal_info: Optional[Dict[str, Any]] = None
    academic_info: Optional[Dict[str, Any]] = None
    academic_bursary_info: Optional[Dict[str, Any]] = None
    employment_info: Optional[Dict[str, Any]] = None
    financial_info: Optional[Dict[str, Any]] = None
    documents: Optional[Dict[str, Any]] = None

class BBBEERecord(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=generate_uuid)
    organization_name: str
    bbbee_level: int
    score: float
    verification_date: datetime
    expiry_date: datetime
    status: str = "active"
    documents: List[str] = []
    created_at: datetime = Field(default_factory=current_time)

class SystemSettings(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = "system_settings"
    smtp_host: Optional[str] = None
    smtp_port: int = 587
    smtp_username: Optional[str] = None
    smtp_password: Optional[str] = None
    smtp_use_tls: bool = True
    smtp_from_email: Optional[str] = None
    smtp_from_name: str = "Sentech Bursary System"
    zoom_api_key: Optional[str] = None
    zoom_api_secret: Optional[str] = None
    zoom_account_id: Optional[str] = None
    teams_tenant_id: Optional[str] = None
    teams_client_id: Optional[str] = None
    teams_client_secret: Optional[str] = None
    teams_webhook_url: Optional[str] = None
    company_name: str = "Sentech"
    company_tagline: Optional[str] = "Bursary Management System"
    company_email: Optional[str] = None
    company_phone: Optional[str] = None
    company_address: Optional[str] = None
    primary_color: str = "#0056B3"
    timezone: str = "Africa/Johannesburg"
    currency: str = "ZAR"
    features_enabled: Dict[str, bool] = {}
    page_settings: Dict[str, Dict[str, Any]] = {}
    created_at: datetime = Field(default_factory=current_time)
    updated_at: datetime = Field(default_factory=current_time)

class SystemSettingsUpdate(BaseModel):
    model_config = ConfigDict(extra="ignore")
    smtp_host: Optional[str] = None
    smtp_port: Optional[int] = None
    smtp_username: Optional[str] = None
    smtp_password: Optional[str] = None
    smtp_use_tls: Optional[bool] = None
    smtp_from_email: Optional[str] = None
    smtp_from_name: Optional[str] = None
    zoom_api_key: Optional[str] = None
    zoom_api_secret: Optional[str] = None
    zoom_account_id: Optional[str] = None
    teams_tenant_id: Optional[str] = None
    teams_client_id: Optional[str] = None
    teams_client_secret: Optional[str] = None
    teams_webhook_url: Optional[str] = None
    company_name: Optional[str] = None
    company_tagline: Optional[str] = None
    company_email: Optional[str] = None
    company_phone: Optional[str] = None
    company_address: Optional[str] = None
    primary_color: Optional[str] = None
    timezone: Optional[str] = None
    currency: Optional[str] = None
    features_enabled: Optional[Dict[str, bool]] = None
    page_settings: Optional[Dict[str, Dict[str, Any]]] = None

class Role(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=generate_uuid)
    name: str
    description: Optional[str] = None
    permissions: Dict[str, List[str]] = {}
    is_system: bool = False
    is_active: bool = True
    created_at: datetime = Field(default_factory=current_time)
    updated_at: datetime = Field(default_factory=current_time)

# ============== PERSONAL DEVELOPMENT PLAN MODEL ==============
class PDPEntry(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=generate_uuid)
    user_id: str
    user_name: Optional[str] = None
    learn_what: str               # What do I need to learn?
    action_plan: str              # What will I do to achieve this?
    resources_support: str        # What resources or support will I need?
    success_criteria: str         # What will my success criteria be?
    target_date: Optional[str] = None       # Target date for completion
    review_date: Optional[str] = None       # Target date for review
    status: str = "not_started"   # not_started | in_progress | completed | overdue
    priority: str = "medium"      # low | medium | high
    category: Optional[str] = None
    notes: Optional[str] = None
    assigned_to: Optional[str] = None       # User ID of assigned person
    assigned_to_name: Optional[str] = None  # Name of assigned person
    assigned_to_email: Optional[str] = None # Email of assigned person
    completed_at: Optional[str] = None
    created_at: datetime = Field(default_factory=current_time)
    updated_at: datetime = Field(default_factory=current_time)

class PDPCreate(BaseModel):
    model_config = ConfigDict(extra="ignore")
    learn_what: str
    action_plan: str
    resources_support: str
    success_criteria: str
    target_date: Optional[str] = None
    review_date: Optional[str] = None
    status: Optional[str] = "not_started"
    priority: Optional[str] = "medium"
    category: Optional[str] = None
    notes: Optional[str] = None
    assigned_to: Optional[str] = None
    assigned_to_name: Optional[str] = None
    assigned_to_email: Optional[str] = None

class PDPUpdate(BaseModel):
    model_config = ConfigDict(extra="ignore")
    learn_what: Optional[str] = None
    action_plan: Optional[str] = None
    resources_support: Optional[str] = None
    success_criteria: Optional[str] = None
    target_date: Optional[str] = None
    review_date: Optional[str] = None
    status: Optional[str] = None
    priority: Optional[str] = None
    category: Optional[str] = None
    notes: Optional[str] = None
    assigned_to: Optional[str] = None
    assigned_to_name: Optional[str] = None
    assigned_to_email: Optional[str] = None

# ============== EVENT MODEL ==============
class Event(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=generate_uuid)
    title: str
    description: Optional[str] = None
    start_date: str
    end_date: Optional[str] = None
    start_time: Optional[str] = None
    end_time: Optional[str] = None
    location: Optional[str] = None
    event_type: str = "event"  # event, meeting, deadline, holiday, reminder
    color: Optional[str] = "#0056B3"
    attendees: List[str] = []
    all_day: bool = False
    recurrence: Optional[str] = None
    created_by: Optional[str] = None
    created_by_name: Optional[str] = None
    created_at: datetime = Field(default_factory=current_time)
    updated_at: datetime = Field(default_factory=current_time)

class EventCreate(BaseModel):
    model_config = ConfigDict(extra="ignore")
    title: str
    description: Optional[str] = None
    start_date: str
    end_date: Optional[str] = None
    start_time: Optional[str] = None
    end_time: Optional[str] = None
    location: Optional[str] = None
    event_type: Optional[str] = "event"
    color: Optional[str] = "#0056B3"
    attendees: Optional[List[str]] = []
    all_day: Optional[bool] = False

class EventUpdate(BaseModel):
    model_config = ConfigDict(extra="ignore")
    title: Optional[str] = None
    description: Optional[str] = None
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    start_time: Optional[str] = None
    end_time: Optional[str] = None
    location: Optional[str] = None
    event_type: Optional[str] = None
    color: Optional[str] = None
    attendees: Optional[List[str]] = None
    all_day: Optional[bool] = None

# ============== DOCUMENT STATUS MODEL ==============
class ApplicationDocumentReupload(BaseModel):
    model_config = ConfigDict(extra="ignore")
    document_type: str
    file_name: str
    file_data: str  # base64 encoded or URL
    notes: Optional[str] = None

class RoleCreate(BaseModel):
    name: str
    description: Optional[str] = None
    permissions: Dict[str, List[str]] = {}

class RoleUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    permissions: Optional[Dict[str, List[str]]] = None
    is_active: Optional[bool] = None

class FAQ(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=generate_uuid)
    question: str
    answer: str
    category: str = "general"
    order: int = 0
    is_published: bool = True
    helpful_count: int = 0
    not_helpful_count: int = 0
    created_at: datetime = Field(default_factory=current_time)
    updated_at: datetime = Field(default_factory=current_time)

class FAQCreate(BaseModel):
    question: str
    answer: str
    category: str = "general"
    order: int = 0
    is_published: bool = True

class FAQUpdate(BaseModel):
    question: Optional[str] = None
    answer: Optional[str] = None
    category: Optional[str] = None
    order: Optional[int] = None
    is_published: Optional[bool] = None

# ============== AUTH HELPERS ==============

def create_access_token(data: dict):
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

def verify_password(plain_password: str, hashed_password: str) -> bool:
    return pwd_context.verify(plain_password, hashed_password)

def get_password_hash(password: str) -> str:
    return pwd_context.hash(password)

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        token = credentials.credentials
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        user_id: str = payload.get("sub")
        if user_id is None:
            raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid token")
    except JWTError:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid token")
    
    user = await db.users.find_one({"id": user_id}, {"_id": 0})
    if user is None:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="User not found")
    return user

async def send_email_notification(to_email: str, subject: str, body_text: str):
    """Send email notification. Falls back to logging if SMTP not configured."""
    smtp_host = os.getenv("SMTP_HOST")
    smtp_port = int(os.getenv("SMTP_PORT", "587"))
    smtp_user = os.getenv("SMTP_USERNAME")
    smtp_pass = os.getenv("SMTP_PASSWORD")
    from_email = os.getenv("SMTP_FROM_EMAIL", "noreply@sentech.co.za")
    from_name = os.getenv("SMTP_FROM_NAME", "Sentech Bursary")

    if not smtp_host or not smtp_user:
        logger.info(f"[EMAIL] To: {to_email} | Subject: {subject} | Body: {body_text[:100]}...")
        return

    try:
        import smtplib
        from email.mime.text import MIMEText
        from email.mime.multipart import MIMEMultipart
        msg = MIMEMultipart()
        msg["From"] = f"{from_name} <{from_email}>"
        msg["To"] = to_email
        msg["Subject"] = subject
        msg.attach(MIMEText(body_text, "html"))
        with smtplib.SMTP(smtp_host, smtp_port) as server:
            server.starttls()
            server.login(smtp_user, smtp_pass)
            server.sendmail(from_email, to_email, msg.as_string())
        logger.info(f"[EMAIL SENT] To: {to_email} | Subject: {subject}")
    except Exception as e:
        logger.error(f"[EMAIL FAILED] To: {to_email} | Error: {e}")

def serialize_doc(doc):
    """Convert datetime fields to ISO format strings"""
    if doc is None:
        return None
    for key, value in doc.items():
        if isinstance(value, datetime):
            doc[key] = value.isoformat()
    return doc

# ============== AUTH ROUTES ==============
# Moved to routers/auth.py



# ============== USERS ROUTES ==============
# Moved to routers/users.py

# ============== TEAMS ROUTES ==============

@api_router.get("/teams")
async def get_teams(current_user: dict = Depends(get_current_user)):
    teams = await db.teams.find({}, {"_id": 0}).to_list(1000)
    return teams

@api_router.get("/teams/{team_id}")
async def get_team(team_id: str, current_user: dict = Depends(get_current_user)):
    team = await db.teams.find_one({"id": team_id}, {"_id": 0})
    if not team:
        raise HTTPException(status_code=404, detail="Team not found")
    return team

@api_router.post("/teams")
async def create_team(team_data: TeamCreate, current_user: dict = Depends(get_current_user)):
    team = Team(**team_data.model_dump())
    team_dict = team.model_dump()
    team_dict['created_at'] = team_dict['created_at'].isoformat()
    team_dict['updated_at'] = team_dict['updated_at'].isoformat()
    await db.teams.insert_one({**team_dict})
    return team

@api_router.put("/teams/{team_id}")
async def update_team(team_id: str, update_data: TeamUpdate, current_user: dict = Depends(get_current_user)):
    team = await db.teams.find_one({"id": team_id}, {"_id": 0})
    if not team:
        raise HTTPException(status_code=404, detail="Team not found")
    
    update_dict = {k: v for k, v in update_data.model_dump().items() if v is not None}
    update_dict["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.teams.update_one({"id": team_id}, {"$set": update_dict})
    return {"message": "Team updated successfully"}

@api_router.delete("/teams/{team_id}")
async def delete_team(team_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.teams.delete_one({"id": team_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Team not found")
    return {"message": "Team deleted successfully"}

@api_router.get("/teams/{team_id}/members")
async def get_team_members(team_id: str, current_user: dict = Depends(get_current_user)):
    team = await db.teams.find_one({"id": team_id}, {"_id": 0})
    if not team:
        raise HTTPException(status_code=404, detail="Team not found")
    
    members = await db.users.find({"team_id": team_id}, {"_id": 0, "password_hash": 0}).to_list(100)
    return members

@api_router.post("/teams/{team_id}/members/{user_id}")
async def add_team_member(team_id: str, user_id: str, current_user: dict = Depends(get_current_user)):
    team = await db.teams.find_one({"id": team_id}, {"_id": 0})
    if not team:
        raise HTTPException(status_code=404, detail="Team not found")
    
    member_ids = team.get("member_ids", [])
    if user_id not in member_ids:
        member_ids.append(user_id)
        await db.teams.update_one({"id": team_id}, {"$set": {"member_ids": member_ids}})
    
    await db.users.update_one({"id": user_id}, {"$set": {"team_id": team_id}})
    return {"message": "Member added successfully"}

@api_router.delete("/teams/{team_id}/members/{user_id}")
async def remove_team_member(team_id: str, user_id: str, current_user: dict = Depends(get_current_user)):
    team = await db.teams.find_one({"id": team_id}, {"_id": 0})
    if not team:
        raise HTTPException(status_code=404, detail="Team not found")
    
    member_ids = team.get("member_ids", [])
    if user_id in member_ids:
        member_ids.remove(user_id)
        await db.teams.update_one({"id": team_id}, {"$set": {"member_ids": member_ids}})
    
    await db.users.update_one({"id": user_id}, {"$set": {"team_id": None}})
    return {"message": "Member removed successfully"}

# ============== MEETINGS ROUTES ==============

@api_router.get("/meetings")
async def get_meetings(
    status: Optional[str] = None,
    meeting_type: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {"$or": [
        {"organizer_id": current_user["id"]},
        {"attendee_ids": current_user["id"]}
    ]}
    if status:
        query["status"] = status
    if meeting_type:
        query["meeting_type"] = meeting_type
    
    meetings = await db.meetings.find(query, {"_id": 0}).to_list(1000)
    return meetings

@api_router.get("/meetings/{meeting_id}")
async def get_meeting(meeting_id: str, current_user: dict = Depends(get_current_user)):
    meeting = await db.meetings.find_one({"id": meeting_id}, {"_id": 0})
    if not meeting:
        raise HTTPException(status_code=404, detail="Meeting not found")
    return meeting

@api_router.post("/meetings")
async def create_meeting(meeting_data: MeetingCreate, current_user: dict = Depends(get_current_user)):
    meeting = Meeting(
        **meeting_data.model_dump(),
        organizer_id=current_user["id"]
    )
    meeting_dict = meeting.model_dump()
    meeting_dict['start_time'] = meeting_dict['start_time'].isoformat()
    meeting_dict['end_time'] = meeting_dict['end_time'].isoformat()
    meeting_dict['created_at'] = meeting_dict['created_at'].isoformat()
    meeting_dict['updated_at'] = meeting_dict['updated_at'].isoformat()
    
    await db.meetings.insert_one({**meeting_dict})
    
    # Auto-create event for this meeting
    event_dict = {
        "id": generate_uuid(),
        "title": meeting_dict.get("title", "Meeting"),
        "description": meeting_dict.get("description", ""),
        "start_date": meeting_dict.get("start_time"),
        "end_date": meeting_dict.get("end_time"),
        "event_type": "meeting",
        "location": meeting_dict.get("location", ""),
        "organizer_id": current_user["id"],
        "attendee_ids": meeting_dict.get("attendee_ids", []),
        "meeting_id": meeting_dict["id"],
        "meeting_link": meeting_dict.get("meeting_link", ""),
        "is_recurring": False,
        "created_by": current_user["id"],
        "created_at": meeting_dict["created_at"],
        "updated_at": meeting_dict["updated_at"],
    }
    await db.events.insert_one({**event_dict})
    
    # Create notification for attendees
    for attendee_id in meeting_dict.get("attendee_ids", []):
        if attendee_id != current_user["id"]:
            notif = {
                "id": generate_uuid(),
                "user_id": attendee_id,
                "type": "meeting_invite",
                "title": "New Meeting Invitation",
                "message": f"You've been invited to: {meeting_dict.get('title', 'Meeting')}",
                "reference_id": meeting_dict["id"],
                "reference_type": "meeting",
                "is_read": False,
                "created_at": datetime.now(timezone.utc).isoformat(),
            }
            await db.notifications.insert_one({**notif})
    
    return meeting

@api_router.put("/meetings/{meeting_id}")
async def update_meeting(meeting_id: str, update_data: MeetingUpdate, current_user: dict = Depends(get_current_user)):
    meeting = await db.meetings.find_one({"id": meeting_id}, {"_id": 0})
    if not meeting:
        raise HTTPException(status_code=404, detail="Meeting not found")
    
    update_dict = {k: v for k, v in update_data.model_dump().items() if v is not None}
    update_dict["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    for field in ['start_time', 'end_time']:
        if field in update_dict and isinstance(update_dict[field], datetime):
            update_dict[field] = update_dict[field].isoformat()
    
    await db.meetings.update_one({"id": meeting_id}, {"$set": update_dict})
    return {"message": "Meeting updated successfully"}

@api_router.delete("/meetings/{meeting_id}")
async def delete_meeting(meeting_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.meetings.delete_one({"id": meeting_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Meeting not found")
    return {"message": "Meeting deleted successfully"}

# ============== NOTES ROUTES ==============

@api_router.get("/notes")
async def get_notes(
    folder_id: Optional[str] = None,
    is_shared: Optional[bool] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {"$or": [
        {"created_by": current_user["id"]},
        {"shared_with": current_user["id"]},
        {"shared_with_teams": current_user.get("team_id")}
    ]}
    if folder_id:
        query["folder_id"] = folder_id
    if is_shared is not None:
        query["is_shared"] = is_shared
    
    notes = await db.notes.find(query, {"_id": 0}).to_list(1000)
    return notes

@api_router.get("/notes/shared")
async def get_shared_notes(current_user: dict = Depends(get_current_user)):
    query = {
        "is_shared": True,
        "$or": [
            {"shared_with": current_user["id"]},
            {"shared_with_teams": current_user.get("team_id")}
        ],
        "created_by": {"$ne": current_user["id"]}
    }
    notes = await db.notes.find(query, {"_id": 0}).to_list(1000)
    return notes

@api_router.get("/notes/folders/list")
async def get_note_folders(current_user: dict = Depends(get_current_user)):
    folders = await db.note_folders.find(
        {"$or": [
            {"created_by": current_user["id"]},
            {"shared_with": current_user["id"]}
        ]},
        {"_id": 0}
    ).to_list(100)
    return folders

@api_router.post("/notes/folders")
async def create_note_folder(folder_data: NoteFolderCreate, current_user: dict = Depends(get_current_user)):
    folder = NoteFolder(
        **folder_data.model_dump(),
        created_by=current_user["id"]
    )
    folder_dict = folder.model_dump()
    folder_dict['created_at'] = folder_dict['created_at'].isoformat()
    folder_dict['updated_at'] = folder_dict['updated_at'].isoformat()
    
    await db.note_folders.insert_one({**folder_dict})
    return folder

@api_router.get("/notes/{note_id}")
async def get_note(note_id: str, current_user: dict = Depends(get_current_user)):
    note = await db.notes.find_one({"id": note_id}, {"_id": 0})
    if not note:
        raise HTTPException(status_code=404, detail="Note not found")
    return note

@api_router.post("/notes")
async def create_note(note_data: NoteCreate, current_user: dict = Depends(get_current_user)):
    note = Note(
        **note_data.model_dump(),
        created_by=current_user["id"]
    )
    note_dict = note.model_dump()
    note_dict['created_at'] = note_dict['created_at'].isoformat()
    note_dict['updated_at'] = note_dict['updated_at'].isoformat()
    
    await db.notes.insert_one({**note_dict})
    return note

@api_router.put("/notes/{note_id}")
async def update_note(note_id: str, update_data: NoteUpdate, current_user: dict = Depends(get_current_user)):
    note = await db.notes.find_one({"id": note_id}, {"_id": 0})
    if not note:
        raise HTTPException(status_code=404, detail="Note not found")
    
    if note["created_by"] != current_user["id"]:
        raise HTTPException(status_code=403, detail="Only the creator can edit this note")
    
    update_dict = {k: v for k, v in update_data.model_dump().items() if v is not None}
    update_dict["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.notes.update_one({"id": note_id}, {"$set": update_dict})
    updated_note = await db.notes.find_one({"id": note_id}, {"_id": 0})
    return updated_note

@api_router.delete("/notes/{note_id}")
async def delete_note(note_id: str, current_user: dict = Depends(get_current_user)):
    note = await db.notes.find_one({"id": note_id}, {"_id": 0})
    if not note:
        raise HTTPException(status_code=404, detail="Note not found")
    
    if note["created_by"] != current_user["id"]:
        raise HTTPException(status_code=403, detail="Only the creator can delete this note")
    
    await db.notes.delete_one({"id": note_id})
    return {"message": "Note deleted successfully"}

@api_router.post("/notes/{note_id}/share")
async def share_note(note_id: str, user_ids: List[str] = [], team_ids: List[str] = [], current_user: dict = Depends(get_current_user)):
    note = await db.notes.find_one({"id": note_id}, {"_id": 0})
    if not note:
        raise HTTPException(status_code=404, detail="Note not found")
    
    if note["created_by"] != current_user["id"]:
        raise HTTPException(status_code=403, detail="Only the creator can share this note")
    
    await db.notes.update_one(
        {"id": note_id},
        {"$set": {
            "is_shared": True,
            "shared_with": user_ids,
            "shared_with_teams": team_ids,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    return {"message": "Note shared successfully"}

# ============== MESSAGES ROUTES ==============

@api_router.get("/messages/conversations")
async def get_conversations(current_user: dict = Depends(get_current_user)):
    conversations = await db.conversations.find(
        {"participant_ids": current_user["id"], "is_archived": False},
        {"_id": 0}
    ).sort("last_message_at", -1).to_list(100)
    
    for conv in conversations:
        participants = await db.users.find(
            {"id": {"$in": conv["participant_ids"]}},
            {"_id": 0, "id": 1, "full_name": 1, "email": 1}
        ).to_list(100)
        conv["participants"] = participants
        conv["unread"] = conv.get("unread_count", {}).get(current_user["id"], 0)
    
    return conversations

@api_router.get("/messages/conversations/{conversation_id}")
async def get_conversation(conversation_id: str, current_user: dict = Depends(get_current_user)):
    conversation = await db.conversations.find_one({"id": conversation_id}, {"_id": 0})
    if not conversation:
        raise HTTPException(status_code=404, detail="Conversation not found")
    
    if current_user["id"] not in conversation["participant_ids"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    participants = await db.users.find(
        {"id": {"$in": conversation["participant_ids"]}},
        {"_id": 0, "id": 1, "full_name": 1, "email": 1}
    ).to_list(100)
    conversation["participants"] = participants
    
    return conversation

@api_router.post("/messages/conversations")
async def create_conversation(conv_data: ConversationCreate, current_user: dict = Depends(get_current_user)):
    if conv_data.type == "direct" and len(conv_data.participant_ids) == 1:
        other_user_id = conv_data.participant_ids[0]
        existing = await db.conversations.find_one({
            "type": "direct",
            "participant_ids": {"$all": [current_user["id"], other_user_id], "$size": 2}
        }, {"_id": 0})
        if existing:
            return existing
    
    participant_ids = list(set(conv_data.participant_ids + [current_user["id"]]))
    
    conversation = Conversation(
        **conv_data.model_dump(exclude={"participant_ids"}),
        participant_ids=participant_ids,
        created_by=current_user["id"]
    )
    conv_dict = conversation.model_dump()
    conv_dict['created_at'] = conv_dict['created_at'].isoformat()
    conv_dict['updated_at'] = conv_dict['updated_at'].isoformat()
    if conv_dict.get('last_message_at'):
        conv_dict['last_message_at'] = conv_dict['last_message_at'].isoformat()
    
    await db.conversations.insert_one({**conv_dict})
    return conversation

@api_router.get("/messages/conversations/{conversation_id}/messages")
async def get_messages(
    conversation_id: str,
    limit: int = 50,
    current_user: dict = Depends(get_current_user)
):
    conversation = await db.conversations.find_one({"id": conversation_id}, {"_id": 0})
    if not conversation:
        raise HTTPException(status_code=404, detail="Conversation not found")
    
    if current_user["id"] not in conversation["participant_ids"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    messages = await db.messages.find(
        {"conversation_id": conversation_id, "is_deleted": False},
        {"_id": 0}
    ).sort("created_at", -1).limit(limit).to_list(limit)
    
    await db.messages.update_many(
        {
            "conversation_id": conversation_id,
            "sender_id": {"$ne": current_user["id"]},
            "is_read": False
        },
        {
            "$set": {"is_read": True},
            "$addToSet": {"read_by": current_user["id"]}
        }
    )
    
    await db.conversations.update_one(
        {"id": conversation_id},
        {"$set": {f"unread_count.{current_user['id']}": 0}}
    )
    
    return list(reversed(messages))

@api_router.post("/messages/conversations/{conversation_id}/messages")
async def send_message(conversation_id: str, message_data: MessageCreate, current_user: dict = Depends(get_current_user)):
    conversation = await db.conversations.find_one({"id": conversation_id}, {"_id": 0})
    if not conversation:
        raise HTTPException(status_code=404, detail="Conversation not found")
    
    if current_user["id"] not in conversation["participant_ids"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    message = Message(
        conversation_id=conversation_id,
        sender_id=current_user["id"],
        content=message_data.content,
        message_type=message_data.message_type,
        attachments=message_data.attachments
    )
    msg_dict = message.model_dump()
    msg_dict['created_at'] = msg_dict['created_at'].isoformat()
    msg_dict['updated_at'] = msg_dict['updated_at'].isoformat()
    
    await db.messages.insert_one({**msg_dict})
    
    unread_update = {}
    for participant_id in conversation["participant_ids"]:
        if participant_id != current_user["id"]:
            current_unread = conversation.get("unread_count", {}).get(participant_id, 0)
            unread_update[f"unread_count.{participant_id}"] = current_unread + 1
    
    await db.conversations.update_one(
        {"id": conversation_id},
        {
            "$set": {
                "last_message_id": message.id,
                "last_message_at": msg_dict['created_at'],
                "updated_at": msg_dict['created_at'],
                **unread_update
            }
        }
    )
    
    # Create notifications for recipients
    await _create_message_notifications(conversation, current_user["id"], current_user.get("full_name", "Unknown"), message_data.content)
    
    return message

@api_router.get("/messages/unread/count")
async def get_unread_count(current_user: dict = Depends(get_current_user)):
    conversations = await db.conversations.find(
        {"participant_ids": current_user["id"]},
        {"_id": 0, "unread_count": 1}
    ).to_list(1000)
    
    total = sum(conv.get("unread_count", {}).get(current_user["id"], 0) for conv in conversations)
    return {"unread_count": total}

# ============== EXPENSES ROUTES ==============

@api_router.get("/expenses")
async def get_expenses(
    status: Optional[str] = None,
    category: Optional[str] = None,
    project_id: Optional[str] = None,
    sponsor_id: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    if "admin" in current_user.get("roles", []):
        query = {}
    else:
        query = {"submitted_by": current_user["id"]}
    
    if status:
        query["status"] = status
    if category:
        query["category"] = category
    if project_id:
        query["project_id"] = project_id
    if sponsor_id:
        query["sponsor_id"] = sponsor_id
    if date_from:
        query.setdefault("date", {})["$gte"] = date_from
    if date_to:
        query.setdefault("date", {})["$lte"] = date_to + "T23:59:59"
    
    expenses = await db.expenses.find(query, {"_id": 0}).sort("date", -1).to_list(1000)
    return expenses

@api_router.get("/expenses/export/excel")
async def export_expenses_excel(
    status: Optional[str] = None,
    category: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    tab: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Export expenses to XLSX with filters"""
    if tab == "bursary" or tab == "training":
        # Export application expenses
        is_admin = "admin" in current_user.get("roles", []) or "super_admin" in current_user.get("roles", [])
        user_query = {} if is_admin else {"user_id": current_user["id"]}
        coll = db.applications if tab == "bursary" else db.training_applications
        apps = await coll.find(
            {**user_query, "additional_expenses": {"$exists": True}}, {"_id": 0}
        ).to_list(10000)
        rows = []
        for app in apps:
            exp = app.get("additional_expenses", {})
            total = sum(exp.get(k, 0) for k in ["flights", "accommodation", "car_hire_or_shuttle", "catering"] if isinstance(exp.get(k), (int, float)))
            if total <= 0:
                continue
            row = {
                "applicant": app.get("applicant_name") or app.get("full_name", ""),
                "status": app.get("status", ""),
                "flights": exp.get("flights", 0),
                "accommodation": exp.get("accommodation", 0),
                "car_hire_shuttle": exp.get("car_hire_or_shuttle", 0),
                "catering": exp.get("catering", 0),
                "total": total,
            }
            if tab == "training":
                row["training_type"] = app.get("training_type") or app.get("service_provider", "")
            rows.append(row)
        excel_stream = generate_excel(rows, title=f"{tab.title()} Expenses")
    else:
        # Export standalone expenses
        query = {} if "admin" in current_user.get("roles", []) else {"submitted_by": current_user["id"]}
        if status:
            query["status"] = status
        if category:
            query["category"] = category
        if date_from:
            query.setdefault("date", {})["$gte"] = date_from
        if date_to:
            query.setdefault("date", {})["$lte"] = date_to + "T23:59:59"
        raw = await db.expenses.find(query, {"_id": 0}).sort("date", -1).to_list(10000)
        rows = []
        for e in raw:
            rows.append({
                "title": e.get("title", ""),
                "category": e.get("category", ""),
                "amount": e.get("amount", 0),
                "date": str(e.get("date", ""))[:10],
                "status": e.get("status", ""),
                "vendor": e.get("vendor", ""),
                "description": e.get("description", ""),
            })
        excel_stream = generate_excel(rows, title="Expenses")

    return StreamingResponse(
        iter([excel_stream.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=expenses_export.xlsx"}
    )

@api_router.get("/expenses/stats")
async def get_expense_stats(current_user: dict = Depends(get_current_user)):
    if "admin" in current_user.get("roles", []):
        query = {}
    else:
        query = {"submitted_by": current_user["id"]}
    
    expenses = await db.expenses.find(query, {"_id": 0}).to_list(10000)
    
    total = sum(e.get("amount", 0) for e in expenses)
    pending = sum(e.get("amount", 0) for e in expenses if e.get("status") == "pending")
    approved = sum(e.get("amount", 0) for e in expenses if e.get("status") == "approved")
    
    by_category = {}
    for e in expenses:
        cat = e.get("category", "other")
        by_category[cat] = by_category.get(cat, 0) + e.get("amount", 0)
    
    return {
        "total": total,
        "pending": pending,
        "approved": approved,
        "by_category": by_category,
        "count": len(expenses)
    }

@api_router.get("/expenses/application-expenses")
async def get_all_application_expenses(current_user: dict = Depends(get_current_user)):
    is_admin = "admin" in current_user.get("roles", []) or "super_admin" in current_user.get("roles", [])
    user_query = {} if is_admin else {"user_id": current_user["id"]}
    
    bursary_apps = await db.applications.find(
        {**user_query, "additional_expenses": {"$exists": True}}, {"_id": 0}
    ).to_list(1000)
    
    training_apps = await db.training_applications.find(
        {**user_query, "additional_expenses": {"$exists": True}}, {"_id": 0}
    ).to_list(1000)
    
    bursary_expenses = []
    for app in bursary_apps:
        exp = app.get("additional_expenses", {})
        total = sum(float(exp.get(k, 0) or 0) for k in ["flights", "accommodation", "car_hire_or_shuttle", "catering"])
        if total > 0:
            user = await db.users.find_one({"id": app["user_id"]}, {"_id": 0, "full_name": 1})
            fi = app.get("financial_info", {})
            requested_amount = fi.get("total_amount") or fi.get("amount_requested") or fi.get("bursary_amount") or 0
            bursary_expenses.append({
                "application_id": app["id"],
                "application_type": "bursary",
                "applicant_name": user.get("full_name", "Unknown") if user else app.get("user_email", "Unknown"),
                "status": app.get("status", "draft"),
                "submitted_at": app.get("submitted_at"),
                "created_at": app.get("created_at"),
                "requested_amount": float(requested_amount) if requested_amount else 0,
                "expenses": exp,
                "total": total,
            })
    
    training_expenses = []
    for app in training_apps:
        exp = app.get("additional_expenses", {})
        total = sum(float(exp.get(k, 0) or 0) for k in ["flights", "accommodation", "car_hire_or_shuttle", "catering"])
        if total > 0:
            user = await db.users.find_one({"id": app["user_id"]}, {"_id": 0, "full_name": 1})
            ti = app.get("training_info", {})
            requested_amount = ti.get("total_amount") or ti.get("amount_requested") or 0
            training_expenses.append({
                "application_id": app["id"],
                "application_type": "training",
                "applicant_name": user.get("full_name", "Unknown") if user else "Unknown",
                "training_type": ti.get("training_type", ""),
                "service_provider": ti.get("service_provider", ""),
                "status": app.get("status", "draft"),
                "submitted_at": app.get("submitted_at"),
                "created_at": app.get("created_at"),
                "requested_amount": float(requested_amount) if requested_amount else 0,
                "expenses": exp,
                "total": total,
            })
    
    return {"bursary": bursary_expenses, "training": training_expenses}

@api_router.get("/expenses/available-applications")
async def get_available_applications_for_expenses(current_user: dict = Depends(get_current_user)):
    """Get all applications (bursary + training) available for expense submission"""
    is_admin = "admin" in current_user.get("roles", []) or "super_admin" in current_user.get("roles", [])
    user_query = {} if is_admin else {"user_id": current_user["id"]}

    bursary_apps = await db.applications.find(
        {**user_query, "status": {"$nin": ["draft"]}}, {"_id": 0}
    ).to_list(1000)
    training_apps = await db.training_applications.find(
        {**user_query, "status": {"$nin": ["draft"]}}, {"_id": 0}
    ).to_list(1000)

    result = []
    for app in bursary_apps:
        user = await db.users.find_one({"id": app["user_id"]}, {"_id": 0, "full_name": 1})
        fi = app.get("financial_info", {})
        requested_amount = fi.get("total_amount") or fi.get("amount_requested") or fi.get("bursary_amount") or 0
        has_expenses = bool(app.get("additional_expenses"))
        result.append({
            "id": app["id"],
            "type": "bursary",
            "applicant_name": user.get("full_name", "Unknown") if user else app.get("user_email", "Unknown"),
            "user_email": app.get("user_email", ""),
            "status": app.get("status", ""),
            "requested_amount": float(requested_amount) if requested_amount else 0,
            "has_expenses": has_expenses,
            "existing_expenses": app.get("additional_expenses") if has_expenses else None,
            "created_at": app.get("created_at"),
        })
    for app in training_apps:
        user = await db.users.find_one({"id": app["user_id"]}, {"_id": 0, "full_name": 1})
        ti = app.get("training_info", {})
        requested_amount = ti.get("total_amount") or ti.get("amount_requested") or 0
        has_expenses = bool(app.get("additional_expenses"))
        result.append({
            "id": app["id"],
            "type": "training",
            "applicant_name": user.get("full_name", "Unknown") if user else app.get("user_email", "Unknown"),
            "user_email": app.get("user_email", ""),
            "status": app.get("status", ""),
            "training_type": ti.get("training_type", ""),
            "service_provider": ti.get("service_provider", ""),
            "requested_amount": float(requested_amount) if requested_amount else 0,
            "has_expenses": has_expenses,
            "existing_expenses": app.get("additional_expenses") if has_expenses else None,
            "created_at": app.get("created_at"),
        })
    return result

@api_router.get("/expenses/{expense_id}")
async def get_expense(expense_id: str, current_user: dict = Depends(get_current_user)):
    expense = await db.expenses.find_one({"id": expense_id}, {"_id": 0})
    if not expense:
        raise HTTPException(status_code=404, detail="Expense not found")
    
    if expense["submitted_by"] != current_user["id"] and "admin" not in current_user.get("roles", []):
        raise HTTPException(status_code=403, detail="Access denied")
    
    return expense

@api_router.post("/expenses")
async def create_expense(expense_data: ExpenseCreate, current_user: dict = Depends(get_current_user)):
    expense = Expense(
        **expense_data.model_dump(),
        submitted_by=current_user["id"]
    )
    expense_dict = expense.model_dump()
    expense_dict['date'] = expense_dict['date'].isoformat()
    expense_dict['created_at'] = expense_dict['created_at'].isoformat()
    expense_dict['updated_at'] = expense_dict['updated_at'].isoformat()
    
    await db.expenses.insert_one({**expense_dict})
    return expense

@api_router.put("/expenses/{expense_id}")
async def update_expense(expense_id: str, update_data: ExpenseUpdate, current_user: dict = Depends(get_current_user)):
    expense = await db.expenses.find_one({"id": expense_id}, {"_id": 0})
    if not expense:
        raise HTTPException(status_code=404, detail="Expense not found")
    
    if expense["submitted_by"] != current_user["id"] and "admin" not in current_user.get("roles", []):
        raise HTTPException(status_code=403, detail="Access denied")
    
    update_dict = {k: v for k, v in update_data.model_dump().items() if v is not None}
    update_dict["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    if 'date' in update_dict and isinstance(update_dict['date'], datetime):
        update_dict['date'] = update_dict['date'].isoformat()
    
    await db.expenses.update_one({"id": expense_id}, {"$set": update_dict})
    return {"message": "Expense updated successfully"}

@api_router.delete("/expenses/{expense_id}")
async def delete_expense(expense_id: str, current_user: dict = Depends(get_current_user)):
    expense = await db.expenses.find_one({"id": expense_id}, {"_id": 0})
    if not expense:
        raise HTTPException(status_code=404, detail="Expense not found")
    
    if expense["submitted_by"] != current_user["id"] and "admin" not in current_user.get("roles", []):
        raise HTTPException(status_code=403, detail="Access denied")
    
    if expense.get("status") != "pending":
        raise HTTPException(status_code=400, detail="Cannot delete approved or rejected expenses")
    
    await db.expenses.delete_one({"id": expense_id})
    return {"message": "Expense deleted successfully"}

@api_router.post("/expenses/{expense_id}/approve")
async def approve_expense(expense_id: str, current_user: dict = Depends(get_current_user)):
    if "admin" not in current_user.get("roles", []):
        raise HTTPException(status_code=403, detail="Only admins can approve expenses")
    
    expense = await db.expenses.find_one({"id": expense_id}, {"_id": 0})
    if not expense:
        raise HTTPException(status_code=404, detail="Expense not found")
    
    await db.expenses.update_one(
        {"id": expense_id},
        {"$set": {
            "status": "approved",
            "approved_by": current_user["id"],
            "approved_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    return {"message": "Expense approved successfully"}

@api_router.post("/expenses/{expense_id}/reject")
async def reject_expense(expense_id: str, reason: str = "", current_user: dict = Depends(get_current_user)):
    if "admin" not in current_user.get("roles", []):
        raise HTTPException(status_code=403, detail="Only admins can reject expenses")
    
    await db.expenses.update_one(
        {"id": expense_id},
        {"$set": {
            "status": "rejected",
            "notes": reason,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    return {"message": "Expense rejected"}

# ============== TICKETS ROUTES ==============

@api_router.get("/tickets")
async def get_tickets(
    status: Optional[str] = None,
    priority: Optional[str] = None,
    category: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    if "admin" in current_user.get("roles", []) or "support" in current_user.get("roles", []):
        query = {}
    else:
        query = {"created_by": current_user["id"]}
    
    if status:
        query["status"] = status
    if priority:
        query["priority"] = priority
    if category:
        query["category"] = category
    
    tickets = await db.tickets.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return tickets

@api_router.get("/tickets/stats")
async def get_ticket_stats(current_user: dict = Depends(get_current_user)):
    total = await db.tickets.count_documents({})
    open_count = await db.tickets.count_documents({"status": "open"})
    in_progress = await db.tickets.count_documents({"status": "in_progress"})
    resolved = await db.tickets.count_documents({"status": "resolved"})
    
    return {
        "total": total,
        "open": open_count,
        "in_progress": in_progress,
        "resolved": resolved
    }

@api_router.get("/tickets/{ticket_id}")
async def get_ticket(ticket_id: str, current_user: dict = Depends(get_current_user)):
    ticket = await db.tickets.find_one({"id": ticket_id}, {"_id": 0})
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    
    if ticket["created_by"] != current_user["id"]:
        if "admin" not in current_user.get("roles", []) and "support" not in current_user.get("roles", []):
            raise HTTPException(status_code=403, detail="Access denied")
    
    return ticket

@api_router.post("/tickets")
async def create_ticket(ticket_data: TicketCreate, current_user: dict = Depends(get_current_user)):
    ticket = Ticket(
        **ticket_data.model_dump(),
        created_by=current_user["id"]
    )
    ticket_dict = ticket.model_dump()
    ticket_dict['created_at'] = ticket_dict['created_at'].isoformat()
    ticket_dict['updated_at'] = ticket_dict['updated_at'].isoformat()
    
    await db.tickets.insert_one({**ticket_dict})
    return ticket

@api_router.put("/tickets/{ticket_id}")
async def update_ticket(ticket_id: str, update_data: TicketUpdate, current_user: dict = Depends(get_current_user)):
    ticket = await db.tickets.find_one({"id": ticket_id}, {"_id": 0})
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    
    if ticket["created_by"] != current_user["id"]:
        if "admin" not in current_user.get("roles", []) and "support" not in current_user.get("roles", []):
            raise HTTPException(status_code=403, detail="Access denied")
    
    update_dict = {k: v for k, v in update_data.model_dump().items() if v is not None}
    update_dict["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    if update_data.status == "resolved" and ticket.get("status") != "resolved":
        update_dict["resolved_at"] = datetime.now(timezone.utc).isoformat()
    
    if update_data.status == "closed" and ticket.get("status") != "closed":
        update_dict["closed_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.tickets.update_one({"id": ticket_id}, {"$set": update_dict})
    
    # Notify ticket creator about status change
    if update_data.status and update_data.status != ticket.get("status") and ticket["created_by"] != current_user["id"]:
        notif = {
            "id": generate_uuid(),
            "user_id": ticket["created_by"],
            "type": "ticket_response",
            "title": f"Ticket Status Updated",
            "message": f"Your ticket '{ticket.get('title', '')}' has been updated to: {update_data.status}",
            "reference_id": ticket_id,
            "reference_type": "ticket",
            "is_read": False,
            "created_at": datetime.now(timezone.utc).isoformat(),
        }
        await db.notifications.insert_one({**notif})
    
    return {"message": "Ticket updated successfully"}

@api_router.delete("/tickets/{ticket_id}")
async def delete_ticket(ticket_id: str, current_user: dict = Depends(get_current_user)):
    ticket = await db.tickets.find_one({"id": ticket_id}, {"_id": 0})
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    
    if ticket["created_by"] != current_user["id"] and "admin" not in current_user.get("roles", []):
        raise HTTPException(status_code=403, detail="Access denied")
    
    await db.tickets.delete_one({"id": ticket_id})
    await db.ticket_comments.delete_many({"ticket_id": ticket_id})
    return {"message": "Ticket deleted successfully"}

@api_router.get("/tickets/{ticket_id}/comments")
async def get_ticket_comments(ticket_id: str, current_user: dict = Depends(get_current_user)):
    ticket = await db.tickets.find_one({"id": ticket_id}, {"_id": 0})
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    
    is_admin_or_support = "admin" in current_user.get("roles", []) or "support" in current_user.get("roles", [])
    if ticket["created_by"] != current_user["id"] and not is_admin_or_support:
        raise HTTPException(status_code=403, detail="Access denied")
    
    query = {"ticket_id": ticket_id}
    if not is_admin_or_support:
        query["is_internal"] = False
    
    comments = await db.ticket_comments.find(query, {"_id": 0}).sort("created_at", 1).to_list(1000)
    
    for comment in comments:
        user = await db.users.find_one({"id": comment["user_id"]}, {"_id": 0, "full_name": 1, "email": 1})
        comment["user"] = user
    
    return comments

@api_router.post("/tickets/{ticket_id}/comments")
async def add_ticket_comment(ticket_id: str, comment_data: TicketCommentCreate, current_user: dict = Depends(get_current_user)):
    ticket = await db.tickets.find_one({"id": ticket_id}, {"_id": 0})
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    
    comment = TicketComment(
        ticket_id=ticket_id,
        user_id=current_user["id"],
        content=comment_data.content,
        is_internal=comment_data.is_internal
    )
    comment_dict = comment.model_dump()
    comment_dict['created_at'] = comment_dict['created_at'].isoformat()
    comment_dict['updated_at'] = comment_dict['updated_at'].isoformat()
    
    await db.ticket_comments.insert_one({**comment_dict})
    
    await db.tickets.update_one(
        {"id": ticket_id},
        {"$set": {"updated_at": comment_dict['created_at']}}
    )
    
    # Notify ticket creator about new comment (if commenter is not the creator and not internal)
    if not comment_data.is_internal and ticket["created_by"] != current_user["id"]:
        notif = {
            "id": generate_uuid(),
            "user_id": ticket["created_by"],
            "type": "ticket_response",
            "title": "New Response on Your Ticket",
            "message": f"New comment on '{ticket.get('title', '')}': {comment_data.content[:80]}",
            "reference_id": ticket_id,
            "reference_type": "ticket",
            "is_read": False,
            "created_at": datetime.now(timezone.utc).isoformat(),
        }
        await db.notifications.insert_one({**notif})
    
    return comment

# ============== FILES ROUTES ==============

@api_router.get("/files")
async def get_files(
    folder_id: Optional[str] = None,
    project_id: Optional[str] = None,
    file_type: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {"$or": [
        {"uploaded_by": current_user["id"]},
        {"shared_with": current_user["id"]},
        {"shared_with_teams": current_user.get("team_id")}
    ]}
    
    if folder_id:
        query["folder_id"] = folder_id
    if project_id:
        query["project_id"] = project_id
    if file_type:
        query["file_type"] = file_type
    
    files = await db.files.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return files

@api_router.get("/files/folders/list")
async def get_folders(
    parent_id: Optional[str] = None,
    project_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {"$or": [
        {"created_by": current_user["id"]},
        {"shared_with": current_user["id"]}
    ]}
    
    if parent_id is not None:
        query["parent_id"] = parent_id
    if project_id:
        query["project_id"] = project_id
    
    folders = await db.folders.find(query, {"_id": 0}).to_list(100)
    return folders

@api_router.post("/files/folders")
async def create_folder(folder_data: FolderCreate, current_user: dict = Depends(get_current_user)):
    folder = Folder(
        **folder_data.model_dump(),
        created_by=current_user["id"]
    )
    folder_dict = folder.model_dump()
    folder_dict['created_at'] = folder_dict['created_at'].isoformat()
    folder_dict['updated_at'] = folder_dict['updated_at'].isoformat()
    
    await db.folders.insert_one({**folder_dict})
    return folder

@api_router.get("/files/{file_id}")
async def get_file(file_id: str, current_user: dict = Depends(get_current_user)):
    file = await db.files.find_one({"id": file_id}, {"_id": 0})
    if not file:
        raise HTTPException(status_code=404, detail="File not found")
    return file

@api_router.post("/files")
async def create_file(file_data: FileCreate, current_user: dict = Depends(get_current_user)):
    file = FileModel(
        **file_data.model_dump(),
        uploaded_by=current_user["id"]
    )
    file_dict = file.model_dump()
    file_dict['created_at'] = file_dict['created_at'].isoformat()
    file_dict['updated_at'] = file_dict['updated_at'].isoformat()
    
    await db.files.insert_one({**file_dict})
    return file

@api_router.put("/files/{file_id}")
async def update_file(file_id: str, update_data: FileUpdate, current_user: dict = Depends(get_current_user)):
    file = await db.files.find_one({"id": file_id}, {"_id": 0})
    if not file:
        raise HTTPException(status_code=404, detail="File not found")
    
    if file["uploaded_by"] != current_user["id"]:
        raise HTTPException(status_code=403, detail="Only the uploader can edit this file")
    
    update_dict = {k: v for k, v in update_data.model_dump().items() if v is not None}
    update_dict["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.files.update_one({"id": file_id}, {"$set": update_dict})
    return {"message": "File updated successfully"}

@api_router.delete("/files/{file_id}")
async def delete_file(file_id: str, current_user: dict = Depends(get_current_user)):
    file = await db.files.find_one({"id": file_id}, {"_id": 0})
    if not file:
        raise HTTPException(status_code=404, detail="File not found")
    
    if file["uploaded_by"] != current_user["id"] and "admin" not in current_user.get("roles", []):
        raise HTTPException(status_code=403, detail="Access denied")
    
    await db.files.delete_one({"id": file_id})
    return {"message": "File deleted successfully"}

@api_router.post("/files/{file_id}/share")
async def share_file(file_id: str, user_ids: List[str] = [], team_ids: List[str] = [], current_user: dict = Depends(get_current_user)):
    file = await db.files.find_one({"id": file_id}, {"_id": 0})
    if not file:
        raise HTTPException(status_code=404, detail="File not found")
    
    if file["uploaded_by"] != current_user["id"]:
        raise HTTPException(status_code=403, detail="Only the uploader can share this file")
    
    await db.files.update_one(
        {"id": file_id},
        {"$set": {
            "is_shared": True,
            "shared_with": user_ids,
            "shared_with_teams": team_ids,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    return {"message": "File shared successfully"}

@api_router.post("/files/upload")
async def upload_file_multipart(
    file: UploadFile = File(...),
    folder_id: Optional[str] = Form(None),
    project_id: Optional[str] = Form(None),
    category: Optional[str] = Form(None),
    current_user: dict = Depends(get_current_user)
):
    """Upload a file using multipart form data instead of base64"""
    import base64
    MAX_SIZE = 20 * 1024 * 1024  # 20MB
    content = await file.read()
    if len(content) > MAX_SIZE:
        raise HTTPException(status_code=413, detail="File too large. Maximum size is 20MB")

    # Store as base64 in MongoDB for backward compatibility
    b64_data = base64.b64encode(content).decode('utf-8')
    file_ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
    content_type = file.content_type or 'application/octet-stream'

    file_doc = {
        "id": generate_uuid(),
        "name": file.filename,
        "file_type": file_ext,
        "content_type": content_type,
        "size": len(content),
        "file_data": f"data:{content_type};base64,{b64_data}",
        "folder_id": folder_id,
        "project_id": project_id,
        "category": category,
        "uploaded_by": current_user["id"],
        "is_shared": False,
        "shared_with": [],
        "shared_with_teams": [],
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.files.insert_one({**file_doc})
    file_doc.pop("file_data", None)
    return file_doc

# ============== TASKS ROUTES ==============

@api_router.get("/tasks")
async def get_tasks(
    status: Optional[str] = None,
    priority: Optional[str] = None,
    assigned_to: Optional[str] = None,
    project_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if status:
        query["status"] = status
    if priority:
        query["priority"] = priority
    if assigned_to:
        query["assigned_to"] = assigned_to
    if project_id:
        query["project_id"] = project_id
    
    tasks = await db.tasks.find(query, {"_id": 0}).to_list(1000)
    return tasks

@api_router.get("/tasks/{task_id}")
async def get_task(task_id: str, current_user: dict = Depends(get_current_user)):
    task = await db.tasks.find_one({"id": task_id}, {"_id": 0})
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    return task

@api_router.post("/tasks")
async def create_task(task_data: TaskCreate, current_user: dict = Depends(get_current_user)):
    task = Task(
        **task_data.model_dump(),
        created_by=current_user["id"]
    )
    
    # Get assignee name
    if task.assigned_to:
        assignee = await db.users.find_one({"id": task.assigned_to}, {"_id": 0, "full_name": 1})
        if assignee:
            task.assignee_name = assignee.get("full_name")
    
    # Get project name
    if task.project_id:
        project = await db.projects.find_one({"id": task.project_id}, {"_id": 0, "name": 1})
        if project:
            task.project_name = project.get("name")
    
    task_dict = task.model_dump()
    task_dict['created_at'] = task_dict['created_at'].isoformat()
    task_dict['updated_at'] = task_dict['updated_at'].isoformat()
    if task_dict.get('due_date'):
        task_dict['due_date'] = task_dict['due_date'].isoformat()
    if task_dict.get('start_date'):
        task_dict['start_date'] = task_dict['start_date'].isoformat()
    
    await db.tasks.insert_one({**task_dict})
    return task

@api_router.put("/tasks/{task_id}")
async def update_task(task_id: str, update_data: TaskUpdate, current_user: dict = Depends(get_current_user)):
    task = await db.tasks.find_one({"id": task_id}, {"_id": 0})
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    
    update_dict = {k: v for k, v in update_data.model_dump().items() if v is not None}
    update_dict["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    # Handle status completion
    if update_data.status == "completed" and task.get("status") != "completed":
        update_dict["completed_at"] = datetime.now(timezone.utc).isoformat()
        update_dict["progress"] = 100
    
    # Convert datetime fields
    for field in ['due_date', 'start_date']:
        if field in update_dict and isinstance(update_dict[field], datetime):
            update_dict[field] = update_dict[field].isoformat()
    
    # Get assignee name if changed
    if update_data.assigned_to:
        assignee = await db.users.find_one({"id": update_data.assigned_to}, {"_id": 0, "full_name": 1})
        if assignee:
            update_dict["assignee_name"] = assignee.get("full_name")
    
    await db.tasks.update_one({"id": task_id}, {"$set": update_dict})
    updated_task = await db.tasks.find_one({"id": task_id}, {"_id": 0})
    return updated_task

@api_router.delete("/tasks/{task_id}")
async def delete_task(task_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.tasks.delete_one({"id": task_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Task not found")
    return {"message": "Task deleted successfully"}

# ============== PROJECTS ROUTES ==============

@api_router.get("/projects")
async def get_projects(
    status: Optional[str] = None,
    project_type: Optional[str] = None,
    client_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if status:
        query["status"] = status
    if project_type:
        query["project_type"] = project_type
    if client_id:
        query["client_id"] = client_id
    
    projects = await db.projects.find(query, {"_id": 0}).to_list(1000)
    return projects

@api_router.get("/projects/{project_id}")
async def get_project(project_id: str, current_user: dict = Depends(get_current_user)):
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    return project

@api_router.post("/projects")
async def create_project(project_data: ProjectCreate, current_user: dict = Depends(get_current_user)):
    project = Project(
        **project_data.model_dump(),
        created_by=current_user["id"]
    )
    
    # Get client name if provided
    if project.client_id:
        client = await db.clients.find_one({"id": project.client_id}, {"_id": 0, "name": 1})
        if client:
            project.client_name = client.get("name")
    
    project_dict = project.model_dump()
    project_dict['created_at'] = project_dict['created_at'].isoformat()
    project_dict['updated_at'] = project_dict['updated_at'].isoformat()
    for field in ['start_date', 'end_date', 'deadline']:
        if project_dict.get(field):
            project_dict[field] = project_dict[field].isoformat()
    
    await db.projects.insert_one({**project_dict})
    return project

@api_router.put("/projects/{project_id}")
async def update_project(project_id: str, update_data: ProjectUpdate, current_user: dict = Depends(get_current_user)):
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    update_dict = {k: v for k, v in update_data.model_dump().items() if v is not None}
    update_dict["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    for field in ['start_date', 'end_date', 'deadline']:
        if field in update_dict and isinstance(update_dict[field], datetime):
            update_dict[field] = update_dict[field].isoformat()
    
    await db.projects.update_one({"id": project_id}, {"$set": update_dict})
    return {"message": "Project updated successfully"}

@api_router.delete("/projects/{project_id}")
async def delete_project(project_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.projects.delete_one({"id": project_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Project not found")
    return {"message": "Project deleted successfully"}

# Clients
@api_router.get("/clients")
async def get_clients(current_user: dict = Depends(get_current_user)):
    clients = await db.clients.find({}, {"_id": 0}).to_list(1000)
    return clients

@api_router.post("/clients")
async def create_client(client_data: ClientCreate, current_user: dict = Depends(get_current_user)):
    client = Client(
        **client_data.model_dump(),
        created_by=current_user["id"]
    )
    client_dict = client.model_dump()
    client_dict['created_at'] = client_dict['created_at'].isoformat()
    client_dict['updated_at'] = client_dict['updated_at'].isoformat()
    
    await db.clients.insert_one({**client_dict})
    return client

# ============== LEADS ROUTES ==============

@api_router.get("/leads")
async def get_leads(
    status: Optional[str] = None,
    source: Optional[str] = None,
    assigned_to: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if status:
        query["status"] = status
    if source:
        query["source"] = source
    if assigned_to:
        query["assigned_to"] = assigned_to
    
    leads = await db.leads.find(query, {"_id": 0}).to_list(1000)
    return leads

@api_router.get("/leads/{lead_id}")
async def get_lead(lead_id: str, current_user: dict = Depends(get_current_user)):
    lead = await db.leads.find_one({"id": lead_id}, {"_id": 0})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    return lead

@api_router.post("/leads")
async def create_lead(lead_data: LeadCreate, current_user: dict = Depends(get_current_user)):
    lead = Lead(
        **lead_data.model_dump(),
        created_by=current_user["id"]
    )
    
    if lead.assigned_to:
        assignee = await db.users.find_one({"id": lead.assigned_to}, {"_id": 0, "full_name": 1})
        if assignee:
            lead.assignee_name = assignee.get("full_name")
    
    lead_dict = lead.model_dump()
    lead_dict['created_at'] = lead_dict['created_at'].isoformat()
    lead_dict['updated_at'] = lead_dict['updated_at'].isoformat()
    for field in ['last_contact_date', 'next_follow_up']:
        if lead_dict.get(field):
            lead_dict[field] = lead_dict[field].isoformat()
    
    await db.leads.insert_one({**lead_dict})
    return lead

@api_router.put("/leads/{lead_id}")
async def update_lead(lead_id: str, update_data: LeadUpdate, current_user: dict = Depends(get_current_user)):
    lead = await db.leads.find_one({"id": lead_id}, {"_id": 0})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    
    update_dict = {k: v for k, v in update_data.model_dump().items() if v is not None}
    update_dict["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    if update_data.assigned_to:
        assignee = await db.users.find_one({"id": update_data.assigned_to}, {"_id": 0, "full_name": 1})
        if assignee:
            update_dict["assignee_name"] = assignee.get("full_name")
    
    await db.leads.update_one({"id": lead_id}, {"$set": update_dict})
    updated_lead = await db.leads.find_one({"id": lead_id}, {"_id": 0})
    return updated_lead

@api_router.delete("/leads/{lead_id}")
async def delete_lead(lead_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.leads.delete_one({"id": lead_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Lead not found")
    return {"message": "Lead deleted successfully"}

# ============== PROSPECTS ROUTES ==============

@api_router.get("/prospects")
async def get_prospects(
    status: Optional[str] = None,
    interest_level: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if status:
        query["status"] = status
    if interest_level:
        query["interest_level"] = interest_level
    
    prospects = await db.prospects.find(query, {"_id": 0}).to_list(1000)
    return prospects

@api_router.post("/prospects")
async def create_prospect(prospect_data: ProspectCreate, current_user: dict = Depends(get_current_user)):
    prospect = Prospect(
        **prospect_data.model_dump(),
        created_by=current_user["id"]
    )
    prospect_dict = prospect.model_dump()
    prospect_dict['created_at'] = prospect_dict['created_at'].isoformat()
    prospect_dict['updated_at'] = prospect_dict['updated_at'].isoformat()
    if prospect_dict.get('next_action_date'):
        prospect_dict['next_action_date'] = prospect_dict['next_action_date'].isoformat()
    
    await db.prospects.insert_one({**prospect_dict})
    return prospect

@api_router.put("/prospects/{prospect_id}")
async def update_prospect(prospect_id: str, update_data: ProspectUpdate, current_user: dict = Depends(get_current_user)):
    prospect = await db.prospects.find_one({"id": prospect_id}, {"_id": 0})
    if not prospect:
        raise HTTPException(status_code=404, detail="Prospect not found")
    
    update_dict = {k: v for k, v in update_data.model_dump().items() if v is not None}
    update_dict["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    if 'next_action_date' in update_dict and isinstance(update_dict['next_action_date'], datetime):
        update_dict['next_action_date'] = update_dict['next_action_date'].isoformat()
    
    await db.prospects.update_one({"id": prospect_id}, {"$set": update_dict})
    return {"message": "Prospect updated successfully"}

@api_router.delete("/prospects/{prospect_id}")
async def delete_prospect(prospect_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.prospects.delete_one({"id": prospect_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Prospect not found")
    return {"message": "Prospect deleted successfully"}

# ============== SPONSORS ROUTES ==============

@api_router.get("/sponsors")
async def get_sponsors(
    status: Optional[str] = None,
    type: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if status:
        query["status"] = status
    if type:
        query["type"] = type
    
    sponsors = await db.sponsors.find(query, {"_id": 0}).to_list(1000)
    return sponsors

@api_router.get("/sponsors/{sponsor_id}")
async def get_sponsor(sponsor_id: str, current_user: dict = Depends(get_current_user)):
    sponsor = await db.sponsors.find_one({"id": sponsor_id}, {"_id": 0})
    if not sponsor:
        raise HTTPException(status_code=404, detail="Sponsor not found")
    return sponsor

@api_router.post("/sponsors")
async def create_sponsor(sponsor_data: SponsorCreate, current_user: dict = Depends(get_current_user)):
    sponsor = Sponsor(
        **sponsor_data.model_dump(),
        created_by=current_user["id"]
    )
    sponsor_dict = sponsor.model_dump()
    sponsor_dict['created_at'] = sponsor_dict['created_at'].isoformat()
    sponsor_dict['updated_at'] = sponsor_dict['updated_at'].isoformat()
    
    await db.sponsors.insert_one({**sponsor_dict})
    return sponsor

@api_router.put("/sponsors/{sponsor_id}")
async def update_sponsor(sponsor_id: str, update_data: SponsorUpdate, current_user: dict = Depends(get_current_user)):
    sponsor = await db.sponsors.find_one({"id": sponsor_id}, {"_id": 0})
    if not sponsor:
        raise HTTPException(status_code=404, detail="Sponsor not found")
    
    update_dict = {k: v for k, v in update_data.model_dump().items() if v is not None}
    update_dict["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.sponsors.update_one({"id": sponsor_id}, {"$set": update_dict})
    return {"message": "Sponsor updated successfully"}

@api_router.delete("/sponsors/{sponsor_id}")
async def delete_sponsor(sponsor_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.sponsors.delete_one({"id": sponsor_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Sponsor not found")
    return {"message": "Sponsor deleted successfully"}

# Sponsor Contacts
@api_router.get("/sponsors/{sponsor_id}/contacts")
async def get_sponsor_contacts(sponsor_id: str, current_user: dict = Depends(get_current_user)):
    contacts = await db.sponsor_contacts.find({"sponsor_id": sponsor_id}, {"_id": 0}).to_list(100)
    return contacts

@api_router.post("/sponsors/contacts")
async def create_sponsor_contact(contact_data: SponsorContactCreate, current_user: dict = Depends(get_current_user)):
    contact = SponsorContact(
        **contact_data.model_dump(),
        created_by=current_user["id"]
    )
    contact_dict = contact.model_dump()
    contact_dict['created_at'] = contact_dict['created_at'].isoformat()
    contact_dict['updated_at'] = contact_dict['updated_at'].isoformat()
    
    await db.sponsor_contacts.insert_one({**contact_dict})
    return contact

# ============== APPLICATIONS ROUTES ==============
# Moved to routers/applications.py

# ============== BBBEE ROUTES ==============

@api_router.get("/bbbee")
async def get_bbbee_records(current_user: dict = Depends(get_current_user)):
    records = await db.bbbee.find({}, {"_id": 0}).to_list(1000)
    return records

@api_router.post("/bbbee")
async def create_bbbee_record(record_data: dict, current_user: dict = Depends(get_current_user)):
    record = BBBEERecord(**record_data)
    record_dict = record.model_dump()
    record_dict['verification_date'] = record_dict['verification_date'].isoformat()
    record_dict['expiry_date'] = record_dict['expiry_date'].isoformat()
    record_dict['created_at'] = record_dict['created_at'].isoformat()
    
    await db.bbbee.insert_one({**record_dict})
    return record

# ============== SETTINGS ROUTES ==============

@api_router.get("/settings")
async def get_settings(current_user: dict = Depends(get_current_user)):
    settings = await db.settings.find_one({"id": "system_settings"}, {"_id": 0})
    if not settings:
        default_settings = SystemSettings()
        settings_dict = default_settings.model_dump()
        settings_dict['created_at'] = settings_dict['created_at'].isoformat()
        settings_dict['updated_at'] = settings_dict['updated_at'].isoformat()
        await db.settings.insert_one({**settings_dict})
        return default_settings
    return settings

@api_router.put("/settings")
async def update_settings(update_data: SystemSettingsUpdate, current_user: dict = Depends(get_current_user)):
    # Allow managers and admins (including super_admin) to update settings
    if not any(r in current_user.get("roles", []) for r in ["admin", "manager", "super_admin"]):
        raise HTTPException(status_code=403, detail="Only admins and managers can update settings")
    
    settings = await db.settings.find_one({"id": "system_settings"}, {"_id": 0})
    if not settings:
        default_settings = SystemSettings()
        settings_dict = default_settings.model_dump()
        update_dict = {k: v for k, v in update_data.model_dump().items() if v is not None}
        settings_dict.update(update_dict)
        settings_dict['created_at'] = settings_dict['created_at'].isoformat()
        settings_dict['updated_at'] = datetime.now(timezone.utc).isoformat()
        await db.settings.insert_one({**settings_dict})
        return {k: v for k, v in settings_dict.items() if k != "_id"}
    
    update_dict = {k: v for k, v in update_data.model_dump().items() if v is not None}
    update_dict["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.settings.update_one({"id": "system_settings"}, {"$set": update_dict})
    updated_settings = await db.settings.find_one({"id": "system_settings"}, {"_id": 0})
    return updated_settings

@api_router.post("/settings/smtp/test")
async def test_smtp(email: str, current_user: dict = Depends(get_current_user)):
    if "admin" not in current_user.get("roles", []):
        raise HTTPException(status_code=403, detail="Only admins can test SMTP")
    
    settings = await db.settings.find_one({"id": "system_settings"}, {"_id": 0})
    if not settings or not settings.get("smtp_host"):
        raise HTTPException(status_code=400, detail="SMTP not configured")
    
    return {"message": f"Test email would be sent to {email}", "status": "simulated"}

@api_router.get("/settings/roles")
async def get_roles(current_user: dict = Depends(get_current_user)):
    roles = await db.roles.find({}, {"_id": 0}).to_list(100)
    
    if not roles:
        default_roles = [
            Role(
                name="admin",
                description="Full system access",
                permissions={
                    "applications": ["create", "read", "update", "delete"],
                    "sponsors": ["create", "read", "update", "delete"],
                    "projects": ["create", "read", "update", "delete"],
                    "tasks": ["create", "read", "update", "delete"],
                    "settings": ["create", "read", "update", "delete"],
                },
                is_system=True
            ),
            Role(
                name="manager",
                description="Team management access",
                permissions={
                    "applications": ["create", "read", "update"],
                    "projects": ["create", "read", "update"],
                    "tasks": ["create", "read", "update", "delete"],
                },
                is_system=True
            ),
            Role(
                name="employee",
                description="Standard employee access",
                permissions={
                    "tasks": ["create", "read", "update"],
                    "notes": ["create", "read", "update", "delete"],
                },
                is_system=True
            ),
        ]
        for role in default_roles:
            role_dict = role.model_dump()
            role_dict['created_at'] = role_dict['created_at'].isoformat()
            role_dict['updated_at'] = role_dict['updated_at'].isoformat()
            await db.roles.insert_one({**role_dict})
        
        roles = await db.roles.find({}, {"_id": 0}).to_list(100)
    
    return roles

@api_router.post("/settings/roles")
async def create_role(role_data: RoleCreate, current_user: dict = Depends(get_current_user)):
    is_admin = "admin" in current_user.get("roles", []) or "super_admin" in current_user.get("roles", [])
    if not is_admin:
        raise HTTPException(status_code=403, detail="Only admins can create roles")
    
    existing = await db.roles.find_one({"name": role_data.name}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Role name already exists")
    
    role = Role(**role_data.model_dump())
    role_dict = role.model_dump()
    role_dict['created_at'] = role_dict['created_at'].isoformat()
    role_dict['updated_at'] = role_dict['updated_at'].isoformat()
    
    await db.roles.insert_one({**role_dict})
    return role

@api_router.put("/settings/roles/{role_id}")
async def update_role(role_id: str, update_data: dict, current_user: dict = Depends(get_current_user)):
    is_admin = "admin" in current_user.get("roles", []) or "super_admin" in current_user.get("roles", [])
    if not is_admin:
        raise HTTPException(status_code=403, detail="Only admins can update roles")
    
    role = await db.roles.find_one({"id": role_id}, {"_id": 0})
    if not role:
        raise HTTPException(status_code=404, detail="Role not found")
    
    update_fields = {}
    if "description" in update_data:
        update_fields["description"] = update_data["description"]
    if "permissions" in update_data:
        update_fields["permissions"] = update_data["permissions"]
    update_fields["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.roles.update_one({"id": role_id}, {"$set": update_fields})
    updated = await db.roles.find_one({"id": role_id}, {"_id": 0})
    return updated

@api_router.delete("/settings/roles/{role_id}")
async def delete_role(role_id: str, current_user: dict = Depends(get_current_user)):
    is_admin = "admin" in current_user.get("roles", []) or "super_admin" in current_user.get("roles", [])
    if not is_admin:
        raise HTTPException(status_code=403, detail="Only admins can delete roles")
    
    role = await db.roles.find_one({"id": role_id}, {"_id": 0})
    if not role:
        raise HTTPException(status_code=404, detail="Role not found")
    if role.get("is_system"):
        raise HTTPException(status_code=400, detail="Cannot delete system roles")
    
    await db.roles.delete_one({"id": role_id})
    return {"message": "Role deleted"}

@api_router.get("/settings/dashboard-preferences")
async def get_dashboard_preferences(current_user: dict = Depends(get_current_user)):
    prefs = await db.dashboard_preferences.find_one({"user_id": current_user["id"]}, {"_id": 0})
    if not prefs:
        return {
            "user_id": current_user["id"],
            "visible_widgets": ["applications", "training", "expenses", "users", "tickets", "tasks", "quick_actions"],
            "layout": "default",
        }
    return prefs

@api_router.put("/settings/dashboard-preferences")
async def update_dashboard_preferences(prefs_data: dict, current_user: dict = Depends(get_current_user)):
    now = datetime.now(timezone.utc).isoformat()
    prefs = {
        "user_id": current_user["id"],
        "visible_widgets": prefs_data.get("visible_widgets", []),
        "layout": prefs_data.get("layout", "default"),
        "updated_at": now,
    }
    await db.dashboard_preferences.update_one(
        {"user_id": current_user["id"]},
        {"$set": prefs},
        upsert=True
    )
    return prefs

@api_router.get("/settings/faqs")
async def get_faqs(
    category: Optional[str] = None,
    published_only: bool = True,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if category:
        query["category"] = category
    if published_only:
        query["is_published"] = True
    
    faqs = await db.faqs.find(query, {"_id": 0}).sort("order", 1).to_list(1000)
    return faqs

@api_router.post("/settings/faqs")
async def create_faq(faq_data: FAQCreate, current_user: dict = Depends(get_current_user)):
    if "admin" not in current_user.get("roles", []):
        raise HTTPException(status_code=403, detail="Only admins can create FAQs")
    
    faq = FAQ(**faq_data.model_dump())
    faq_dict = faq.model_dump()
    faq_dict['created_at'] = faq_dict['created_at'].isoformat()
    faq_dict['updated_at'] = faq_dict['updated_at'].isoformat()
    
    await db.faqs.insert_one({**faq_dict})
    return faq

@api_router.put("/settings/faqs/{faq_id}")
async def update_faq(faq_id: str, update_data: FAQUpdate, current_user: dict = Depends(get_current_user)):
    if "admin" not in current_user.get("roles", []):
        raise HTTPException(status_code=403, detail="Only admins can update FAQs")
    
    faq = await db.faqs.find_one({"id": faq_id}, {"_id": 0})
    if not faq:
        raise HTTPException(status_code=404, detail="FAQ not found")
    
    update_dict = {k: v for k, v in update_data.model_dump().items() if v is not None}
    update_dict["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.faqs.update_one({"id": faq_id}, {"$set": update_dict})
    return {"message": "FAQ updated successfully"}

@api_router.delete("/settings/faqs/{faq_id}")
async def delete_faq(faq_id: str, current_user: dict = Depends(get_current_user)):
    if "admin" not in current_user.get("roles", []):
        raise HTTPException(status_code=403, detail="Only admins can delete FAQs")
    
    result = await db.faqs.delete_one({"id": faq_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="FAQ not found")
    return {"message": "FAQ deleted successfully"}

# ============== REPORTS ROUTES ==============
# /reports/dashboard moved to routers/reports.py

@api_router.get("/reports/export/{report_type}")
async def export_report(
    report_type: str,
    format: str = "json",
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    date_query = {}
    if date_from:
        date_query["$gte"] = date_from
    if date_to:
        date_query["$lte"] = date_to + "T23:59:59"

    query = {}
    if date_query:
        query["created_at"] = date_query

    if report_type == "applications":
        data = await db.applications.find(query, {"_id": 0}).to_list(10000)
    elif report_type == "expenses":
        eq = {}
        if date_query:
            eq["date"] = date_query
        data = await db.expenses.find(eq, {"_id": 0}).to_list(10000)
    elif report_type == "tasks":
        data = await db.tasks.find(query, {"_id": 0}).to_list(10000)
    elif report_type == "tickets":
        data = await db.tickets.find(query, {"_id": 0}).to_list(10000)
    elif report_type == "users":
        data = await db.users.find(query, {"_id": 0, "password_hash": 0}).to_list(10000)
    elif report_type == "training_applications":
        data = await db.training_applications.find(query, {"_id": 0}).to_list(10000)
    else:
        raise HTTPException(status_code=400, detail="Invalid report type")
    
    for item in data:
        for key, value in item.items():
            if isinstance(value, datetime):
                item[key] = value.isoformat()
    
    if format == "excel":
        title = report_type.replace("_", " ").title()
        excel_stream = generate_excel(data, title=title)
        return StreamingResponse(
            iter([excel_stream.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={report_type}_report.xlsx"}
        )
    elif format == "pdf":
        title = f"{report_type.replace('_', ' ').title()} Report"
        pdf_stream = generate_pdf(data, title=title)
        return StreamingResponse(
            iter([pdf_stream.getvalue()]),
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={report_type}_report.pdf"}
        )
    elif format == "csv":
        if not data:
            return StreamingResponse(
                io.StringIO("No data"),
                media_type="text/csv",
                headers={"Content-Disposition": f"attachment; filename={report_type}_report.csv"}
            )
        import csv
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=data[0].keys())
        writer.writeheader()
        writer.writerows(data)
        return StreamingResponse(
            io.StringIO(output.getvalue()),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={report_type}_report.csv"}
        )
    else:
        return StreamingResponse(
            io.StringIO(json.dumps(data, indent=2, default=str)),
            media_type="application/json",
            headers={"Content-Disposition": f"attachment; filename={report_type}_report.json"}
        )

# ============== EXPORT HELPERS ==============

def generate_excel(data: list, title: str = "Export") -> io.BytesIO:
    """Generate an Excel file from a list of dicts"""
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    if not data:
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    headers = list(data[0].keys())
    header_fill = PatternFill(start_color="0056B3", end_color="0056B3", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header.replace("_", " ").title())
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    alt_fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, header in enumerate(headers, 1):
            val = row_data.get(header, "")
            if isinstance(val, list):
                val = ", ".join(str(v) for v in val)
            elif isinstance(val, dict):
                val = str(val)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    for col_idx, header in enumerate(headers, 1):
        max_len = max(len(str(header)), max((len(str(row_data.get(header, ""))) for row_data in data), default=0))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_pdf(data: list, title: str = "Export") -> io.BytesIO:
    """Generate a PDF file from a list of dicts"""
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4, rightMargin=0.5*inch, leftMargin=0.5*inch,
                            topMargin=0.5*inch, bottomMargin=0.5*inch, title=title)
    elements = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle('title', parent=styles['Heading1'], fontSize=16,
                                  textColor=colors.HexColor('#0056B3'), spaceAfter=12, alignment=1)
    elements.append(Paragraph(title, title_style))
    elements.append(Spacer(1, 0.1*inch))
    meta = f"<font size=9>Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')} &nbsp;&nbsp; Total: {len(data)} records</font>"
    elements.append(Paragraph(meta, styles['Normal']))
    elements.append(Spacer(1, 0.15*inch))

    if not data:
        elements.append(Paragraph("No data available.", styles['Normal']))
        doc.build(elements)
        output.seek(0)
        return output

    headers = list(data[0].keys())
    display_headers = [h.replace("_", " ").title() for h in headers]
    table_data = [display_headers]
    for row in data:
        r = []
        for h in headers:
            val = row.get(h, "")
            if isinstance(val, list):
                val = ", ".join(str(v) for v in val)
            elif isinstance(val, dict):
                val = str(val)
            r.append(str(val)[:80])
        table_data.append(r)

    col_width = (A4[0] - inch) / len(headers)
    table = Table(table_data, colWidths=[col_width]*len(headers), repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0056B3')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#E8F0FE')]),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('PADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(table)
    doc.build(elements)
    output.seek(0)
    return output


# ============== TASKS EXPORT ROUTES ==============

@api_router.get("/tasks/export/excel")
async def export_tasks_excel(
    status: Optional[str] = None,
    priority: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if status:
        query["status"] = status
    if priority:
        query["priority"] = priority

    tasks = await db.tasks.find(query, {"_id": 0}).to_list(10000)

    if date_from:
        tasks = [t for t in tasks if t.get("due_date") and t["due_date"] >= date_from]
    if date_to:
        tasks = [t for t in tasks if t.get("due_date") and t["due_date"] <= date_to + "T23:59:59"]

    export_data = [{
        "Title": t.get("title", ""),
        "Status": t.get("status", "").replace("_", " ").title(),
        "Priority": t.get("priority", "").title(),
        "Assignee": t.get("assignee_name", ""),
        "Project": t.get("project_name", ""),
        "Due Date": t.get("due_date", "")[:10] if t.get("due_date") else "",
        "Progress": f"{t.get('progress', 0)}%",
        "Tags": ", ".join(t.get("tags", [])),
    } for t in tasks]

    excel_stream = generate_excel(export_data, title="Tasks")
    return StreamingResponse(
        iter([excel_stream.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=tasks_export.xlsx"}
    )


@api_router.get("/tasks/export/pdf")
async def export_tasks_pdf(
    status: Optional[str] = None,
    priority: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if status:
        query["status"] = status
    if priority:
        query["priority"] = priority

    tasks = await db.tasks.find(query, {"_id": 0}).to_list(10000)

    if date_from:
        tasks = [t for t in tasks if t.get("due_date") and t["due_date"] >= date_from]
    if date_to:
        tasks = [t for t in tasks if t.get("due_date") and t["due_date"] <= date_to + "T23:59:59"]

    export_data = [{
        "Title": t.get("title", ""),
        "Status": t.get("status", "").replace("_", " ").title(),
        "Priority": t.get("priority", "").title(),
        "Assignee": t.get("assignee_name", ""),
        "Due Date": t.get("due_date", "")[:10] if t.get("due_date") else "",
        "Progress": f"{t.get('progress', 0)}%",
    } for t in tasks]

    pdf_stream = generate_pdf(export_data, title="Tasks Report")
    return StreamingResponse(
        iter([pdf_stream.getvalue()]),
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=tasks_export.pdf"}
    )


# ============== EVENTS ROUTES ==============

@api_router.get("/events")
async def get_events(
    event_type: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if event_type:
        query["event_type"] = event_type
    if start_date:
        query["start_date"] = {"$gte": start_date}
    if end_date:
        if "start_date" in query:
            query["start_date"]["$lte"] = end_date
        else:
            query["start_date"] = {"$lte": end_date}
    events = await db.events.find(query, {"_id": 0}).sort("start_date", 1).to_list(1000)
    return events

@api_router.get("/events/{event_id}")
async def get_event(event_id: str, current_user: dict = Depends(get_current_user)):
    event = await db.events.find_one({"id": event_id}, {"_id": 0})
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")
    return event

@api_router.post("/events")
async def create_event(event_data: EventCreate, current_user: dict = Depends(get_current_user)):
    event = Event(
        **event_data.model_dump(),
        created_by=current_user["id"],
        created_by_name=current_user.get("full_name", "")
    )
    event_dict = event.model_dump()
    event_dict['created_at'] = event_dict['created_at'].isoformat()
    event_dict['updated_at'] = event_dict['updated_at'].isoformat()
    await db.events.insert_one({**event_dict})
    return event

@api_router.put("/events/{event_id}")
async def update_event(event_id: str, update_data: EventUpdate, current_user: dict = Depends(get_current_user)):
    event = await db.events.find_one({"id": event_id}, {"_id": 0})
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")
    update_dict = {k: v for k, v in update_data.model_dump().items() if v is not None}
    update_dict["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.events.update_one({"id": event_id}, {"$set": update_dict})
    updated = await db.events.find_one({"id": event_id}, {"_id": 0})
    return updated

@api_router.delete("/events/{event_id}")
async def delete_event(event_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.events.delete_one({"id": event_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Event not found")
    return {"message": "Event deleted successfully"}

# ============== USER DOCUMENTS ROUTES ==============
# Moved to routers/users.py

# ============== PERSONAL DEVELOPMENT PLAN ROUTES ==============

@api_router.get("/pdp")
async def get_pdp_entries(
    status: Optional[str] = None,
    priority: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    # Admins can see all; students see only their own
    query = {}
    if "admin" not in current_user.get("roles", []):
        query["user_id"] = current_user["id"]
    if status:
        query["status"] = status
    if priority:
        query["priority"] = priority
    entries = await db.pdp.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return entries

@api_router.get("/pdp/{entry_id}")
async def get_pdp_entry(entry_id: str, current_user: dict = Depends(get_current_user)):
    entry = await db.pdp.find_one({"id": entry_id}, {"_id": 0})
    if not entry:
        raise HTTPException(status_code=404, detail="PDP entry not found")
    if "admin" not in current_user.get("roles", []) and entry.get("user_id") != current_user["id"]:
        raise HTTPException(status_code=403, detail="Access denied")
    return entry

@api_router.post("/pdp")
async def create_pdp_entry(entry_data: PDPCreate, current_user: dict = Depends(get_current_user)):
    entry = PDPEntry(
        **entry_data.model_dump(),
        user_id=current_user["id"],
        user_name=current_user.get("full_name", "")
    )
    entry_dict = entry.model_dump()
    entry_dict["created_at"] = entry_dict["created_at"].isoformat()
    entry_dict["updated_at"] = entry_dict["updated_at"].isoformat()
    await db.pdp.insert_one({**entry_dict})
    return entry

@api_router.put("/pdp/{entry_id}")
async def update_pdp_entry(entry_id: str, update_data: PDPUpdate, current_user: dict = Depends(get_current_user)):
    entry = await db.pdp.find_one({"id": entry_id}, {"_id": 0})
    if not entry:
        raise HTTPException(status_code=404, detail="PDP entry not found")
    if "admin" not in current_user.get("roles", []) and entry.get("user_id") != current_user["id"]:
        raise HTTPException(status_code=403, detail="Access denied")
    update_dict = {k: v for k, v in update_data.model_dump().items() if v is not None}
    update_dict["updated_at"] = datetime.now(timezone.utc).isoformat()
    if update_dict.get("status") == "completed" and not entry.get("completed_at"):
        update_dict["completed_at"] = datetime.now(timezone.utc).isoformat()
    await db.pdp.update_one({"id": entry_id}, {"$set": update_dict})
    updated = await db.pdp.find_one({"id": entry_id}, {"_id": 0})
    return updated

@api_router.delete("/pdp/{entry_id}")
async def delete_pdp_entry(entry_id: str, current_user: dict = Depends(get_current_user)):
    entry = await db.pdp.find_one({"id": entry_id}, {"_id": 0})
    if not entry:
        raise HTTPException(status_code=404, detail="PDP entry not found")
    if "admin" not in current_user.get("roles", []) and entry.get("user_id") != current_user["id"]:
        raise HTTPException(status_code=403, detail="Access denied")
    await db.pdp.delete_one({"id": entry_id})
    return {"message": "PDP entry deleted successfully"}

# ============== DIVISIONS ROUTES ==============

@api_router.get("/divisions")
async def get_divisions(current_user: dict = Depends(get_current_user)):
    divisions = await db.divisions.find({}, {"_id": 0}).to_list(100)
    return divisions

@api_router.post("/divisions")
async def create_division(division_data: dict, current_user: dict = Depends(get_current_user)):
    if "admin" not in current_user.get("roles", []) and "super_admin" not in current_user.get("roles", []):
        raise HTTPException(status_code=403, detail="Only admins can create divisions")
    
    division = {
        "id": generate_uuid(),
        "name": division_data.get("name"),
        "description": division_data.get("description", ""),
        "is_active": True,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.divisions.insert_one({**division})
    return division

@api_router.put("/divisions/{division_id}")
async def update_division(division_id: str, division_data: dict, current_user: dict = Depends(get_current_user)):
    if "admin" not in current_user.get("roles", []) and "super_admin" not in current_user.get("roles", []):
        raise HTTPException(status_code=403, detail="Only admins can update divisions")
    
    update_dict = {k: v for k, v in division_data.items() if v is not None}
    update_dict["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.divisions.update_one({"id": division_id}, {"$set": update_dict})
    updated = await db.divisions.find_one({"id": division_id}, {"_id": 0})
    return updated

@api_router.delete("/divisions/{division_id}")
async def delete_division(division_id: str, current_user: dict = Depends(get_current_user)):
    if "admin" not in current_user.get("roles", []) and "super_admin" not in current_user.get("roles", []):
        raise HTTPException(status_code=403, detail="Only admins can delete divisions")
    
    await db.divisions.delete_one({"id": division_id})
    return {"message": "Division deleted successfully"}

# ============== DEPARTMENTS ROUTES ==============

@api_router.get("/departments")
async def get_departments(division_id: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    query = {}
    if division_id:
        query["division_id"] = division_id
    departments = await db.departments.find(query, {"_id": 0}).to_list(500)
    return departments

@api_router.post("/departments")
async def create_department(dept_data: dict, current_user: dict = Depends(get_current_user)):
    if "admin" not in current_user.get("roles", []) and "super_admin" not in current_user.get("roles", []):
        raise HTTPException(status_code=403, detail="Only admins can create departments")
    
    department = {
        "id": generate_uuid(),
        "name": dept_data.get("name"),
        "division_id": dept_data.get("division_id"),
        "description": dept_data.get("description", ""),
        "is_active": True,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.departments.insert_one({**department})
    return department

@api_router.put("/departments/{dept_id}")
async def update_department(dept_id: str, dept_data: dict, current_user: dict = Depends(get_current_user)):
    if "admin" not in current_user.get("roles", []) and "super_admin" not in current_user.get("roles", []):
        raise HTTPException(status_code=403, detail="Only admins can update departments")
    
    update_dict = {k: v for k, v in dept_data.items() if v is not None}
    update_dict["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.departments.update_one({"id": dept_id}, {"$set": update_dict})
    updated = await db.departments.find_one({"id": dept_id}, {"_id": 0})
    return updated

@api_router.delete("/departments/{dept_id}")
async def delete_department(dept_id: str, current_user: dict = Depends(get_current_user)):
    if "admin" not in current_user.get("roles", []) and "super_admin" not in current_user.get("roles", []):
        raise HTTPException(status_code=403, detail="Only admins can delete departments")
    
    await db.departments.delete_one({"id": dept_id})
    return {"message": "Department deleted successfully"}

# ============== USER IMPORT ROUTES ==============
# Moved to routers/users.py

# ============== DIVISION GROUPS ROUTES ==============

@api_router.get("/division-groups")
async def get_division_groups(current_user: dict = Depends(get_current_user)):
    divisions = await db.divisions.find({}, {"_id": 0}).to_list(100)
    users = await db.users.find({}, {"_id": 0, "password_hash": 0}).to_list(1000)
    group_configs = await db.division_groups.find({}, {"_id": 0}).to_list(100)
    config_map = {gc["division_name"]: gc for gc in group_configs}
    all_subgroups = await db.subgroups.find({}, {"_id": 0}).to_list(500)
    user_map = {u["id"]: u for u in users}
    
    groups = []
    for div in divisions:
        div_name = div["name"]
        all_div_members = [u for u in users if u.get("division") == div_name]
        config = config_map.get(div_name, {})
        div_subgroups = [sg for sg in all_subgroups if sg.get("division_name") == div_name]
        # Collect all user IDs that are in any subgroup of this division
        subgroup_member_ids = set()
        now_str = datetime.now(timezone.utc).isoformat()
        for sg in div_subgroups:
            sg["members"] = [user_map[uid] for uid in sg.get("member_user_ids", []) if uid in user_map]
            sg["leader"] = user_map.get(sg.get("leader_id"))
            # Resolve temp leader
            temp_end = sg.get("temp_leader_end")
            if temp_end and temp_end > now_str and sg.get("temp_leader_id"):
                sg["temp_leader"] = user_map.get(sg.get("temp_leader_id"))
                sg["temp_leader_active"] = True
            else:
                sg["temp_leader"] = None
                sg["temp_leader_active"] = False
            for uid in sg.get("member_user_ids", []):
                subgroup_member_ids.add(uid)
        # Exclude subgroup members from main member list
        main_members = [u for u in all_div_members if u["id"] not in subgroup_member_ids]
        groups.append({
            "division_id": div["id"],
            "division_name": div_name,
            "description": div.get("description", ""),
            "leader_id": config.get("leader_id"),
            "leader": next((u for u in all_div_members if u.get("id") == config.get("leader_id")), None),
            "members": main_members,
            "member_count": len(all_div_members),
            "subgroups": div_subgroups,
        })
    
    groups.sort(key=lambda g: g["division_name"])
    return groups

@api_router.get("/division-groups/{division_name}")
async def get_division_group(division_name: str, current_user: dict = Depends(get_current_user)):
    division = await db.divisions.find_one({"name": division_name}, {"_id": 0})
    if not division:
        raise HTTPException(status_code=404, detail="Division not found")
    
    all_div_members = await db.users.find({"division": division_name}, {"_id": 0, "password_hash": 0}).to_list(200)
    config = await db.division_groups.find_one({"division_name": division_name}, {"_id": 0})
    leader_id = config.get("leader_id") if config else None
    all_users = await db.users.find({}, {"_id": 0, "password_hash": 0}).to_list(1000)
    all_user_map = {u["id"]: u for u in all_users}
    subgroups = await db.subgroups.find({"division_name": division_name}, {"_id": 0}).to_list(100)
    subgroup_member_ids = set()
    now_str = datetime.now(timezone.utc).isoformat()
    for sg in subgroups:
        sg["members"] = [all_user_map[uid] for uid in sg.get("member_user_ids", []) if uid in all_user_map]
        sg["leader"] = all_user_map.get(sg.get("leader_id"))
        temp_end = sg.get("temp_leader_end")
        if temp_end and temp_end > now_str and sg.get("temp_leader_id"):
            sg["temp_leader"] = all_user_map.get(sg.get("temp_leader_id"))
            sg["temp_leader_active"] = True
        else:
            sg["temp_leader"] = None
            sg["temp_leader_active"] = False
        for uid in sg.get("member_user_ids", []):
            subgroup_member_ids.add(uid)
    main_members = [u for u in all_div_members if u["id"] not in subgroup_member_ids]
    
    return {
        "division_id": division["id"],
        "division_name": division_name,
        "description": division.get("description", ""),
        "leader_id": leader_id,
        "leader": next((u for u in all_div_members if u.get("id") == leader_id), None),
        "members": main_members,
        "member_count": len(all_div_members),
        "subgroups": subgroups,
    }

@api_router.put("/division-groups/{division_name}/leader")
async def set_division_group_leader(division_name: str, body: dict, current_user: dict = Depends(get_current_user)):
    if "admin" not in current_user.get("roles", []) and "super_admin" not in current_user.get("roles", []):
        raise HTTPException(status_code=403, detail="Only admins can set group leaders")
    
    division = await db.divisions.find_one({"name": division_name}, {"_id": 0})
    if not division:
        raise HTTPException(status_code=404, detail="Division not found")
    
    leader_id = body.get("leader_id")
    if leader_id:
        user = await db.users.find_one({"id": leader_id}, {"_id": 0})
        if not user:
            raise HTTPException(status_code=404, detail="User not found")
    
    existing = await db.division_groups.find_one({"division_name": division_name})
    if existing:
        await db.division_groups.update_one(
            {"division_name": division_name},
            {"$set": {"leader_id": leader_id, "updated_at": datetime.now(timezone.utc).isoformat()}}
        )
    else:
        await db.division_groups.insert_one({
            "id": generate_uuid(),
            "division_name": division_name,
            "leader_id": leader_id,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat(),
        })
    
    return {"message": f"Leader set for {division_name}", "leader_id": leader_id}

@api_router.post("/division-groups/{division_name}/members/{user_id}")
async def add_member_to_division_group(division_name: str, user_id: str, current_user: dict = Depends(get_current_user)):
    if "admin" not in current_user.get("roles", []) and "super_admin" not in current_user.get("roles", []):
        raise HTTPException(status_code=403, detail="Only admins can manage group members")
    
    division = await db.divisions.find_one({"name": division_name}, {"_id": 0})
    if not division:
        raise HTTPException(status_code=404, detail="Division not found")
    
    user = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    await db.users.update_one({"id": user_id}, {"$set": {"division": division_name, "updated_at": datetime.now(timezone.utc).isoformat()}})
    return {"message": f"User added to {division_name}"}

@api_router.delete("/division-groups/{division_name}/members/{user_id}")
async def remove_member_from_division_group(division_name: str, user_id: str, current_user: dict = Depends(get_current_user)):
    if "admin" not in current_user.get("roles", []) and "super_admin" not in current_user.get("roles", []):
        raise HTTPException(status_code=403, detail="Only admins can manage group members")
    
    await db.users.update_one({"id": user_id}, {"$set": {"division": None, "updated_at": datetime.now(timezone.utc).isoformat()}})
    
    config = await db.division_groups.find_one({"division_name": division_name})
    if config and config.get("leader_id") == user_id:
        await db.division_groups.update_one(
            {"division_name": division_name},
            {"$set": {"leader_id": None, "updated_at": datetime.now(timezone.utc).isoformat()}}
        )
    
    return {"message": f"User removed from {division_name}"}

# ============== SUBGROUPS ROUTES ==============

def _check_admin_or_tech_support(current_user: dict):
    roles = current_user.get("roles", [])
    division = current_user.get("division", "")
    if "admin" in roles or "super_admin" in roles or division == "Technical Support":
        return True
    return False

@api_router.get("/division-groups/{division_name}/subgroups")
async def get_subgroups(division_name: str, current_user: dict = Depends(get_current_user)):
    subgroups = await db.subgroups.find({"division_name": division_name}, {"_id": 0}).to_list(100)
    all_users = await db.users.find({}, {"_id": 0, "password_hash": 0}).to_list(1000)
    user_map = {u["id"]: u for u in all_users}
    now_str = datetime.now(timezone.utc).isoformat()
    for sg in subgroups:
        sg["members"] = [user_map[uid] for uid in sg.get("member_user_ids", []) if uid in user_map]
        sg["leader"] = user_map.get(sg.get("leader_id"))
        temp_end = sg.get("temp_leader_end")
        if temp_end and temp_end > now_str and sg.get("temp_leader_id"):
            sg["temp_leader"] = user_map.get(sg.get("temp_leader_id"))
            sg["temp_leader_active"] = True
        else:
            sg["temp_leader"] = None
            sg["temp_leader_active"] = False
    return subgroups

@api_router.post("/division-groups/{division_name}/subgroups")
async def create_subgroup(division_name: str, body: dict, current_user: dict = Depends(get_current_user)):
    if not _check_admin_or_tech_support(current_user):
        raise HTTPException(status_code=403, detail="Only admins or Technical Support can manage subgroups")
    
    division = await db.divisions.find_one({"name": division_name}, {"_id": 0})
    if not division:
        raise HTTPException(status_code=404, detail="Division not found")
    
    subgroup = {
        "id": generate_uuid(),
        "name": body.get("name", "New Subgroup"),
        "division_name": division_name,
        "leader_id": body.get("leader_id"),
        "member_user_ids": body.get("member_user_ids", []),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.subgroups.insert_one({**subgroup})
    return subgroup

@api_router.put("/subgroups/{subgroup_id}")
async def update_subgroup(subgroup_id: str, body: dict, current_user: dict = Depends(get_current_user)):
    if not _check_admin_or_tech_support(current_user):
        raise HTTPException(status_code=403, detail="Only admins or Technical Support can manage subgroups")
    
    sg = await db.subgroups.find_one({"id": subgroup_id}, {"_id": 0})
    if not sg:
        raise HTTPException(status_code=404, detail="Subgroup not found")
    
    update_dict = {}
    if "name" in body:
        update_dict["name"] = body["name"]
    if "leader_id" in body:
        update_dict["leader_id"] = body["leader_id"]
    update_dict["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.subgroups.update_one({"id": subgroup_id}, {"$set": update_dict})
    updated = await db.subgroups.find_one({"id": subgroup_id}, {"_id": 0})
    return updated

@api_router.post("/subgroups/{subgroup_id}/members/{user_id}")
async def add_subgroup_member(subgroup_id: str, user_id: str, current_user: dict = Depends(get_current_user)):
    if not _check_admin_or_tech_support(current_user):
        raise HTTPException(status_code=403, detail="Only admins or Technical Support can manage subgroups")
    
    sg = await db.subgroups.find_one({"id": subgroup_id}, {"_id": 0})
    if not sg:
        raise HTTPException(status_code=404, detail="Subgroup not found")
    
    member_ids = sg.get("member_user_ids", [])
    if user_id not in member_ids:
        member_ids.append(user_id)
        await db.subgroups.update_one({"id": subgroup_id}, {"$set": {"member_user_ids": member_ids, "updated_at": datetime.now(timezone.utc).isoformat()}})
    
    return {"message": "Member added to subgroup"}

@api_router.delete("/subgroups/{subgroup_id}/members/{user_id}")
async def remove_subgroup_member(subgroup_id: str, user_id: str, current_user: dict = Depends(get_current_user)):
    if not _check_admin_or_tech_support(current_user):
        raise HTTPException(status_code=403, detail="Only admins or Technical Support can manage subgroups")
    
    sg = await db.subgroups.find_one({"id": subgroup_id}, {"_id": 0})
    if not sg:
        raise HTTPException(status_code=404, detail="Subgroup not found")
    
    member_ids = sg.get("member_user_ids", [])
    if user_id in member_ids:
        member_ids.remove(user_id)
        await db.subgroups.update_one({"id": subgroup_id}, {"$set": {"member_user_ids": member_ids, "updated_at": datetime.now(timezone.utc).isoformat()}})
    
    if sg.get("leader_id") == user_id:
        await db.subgroups.update_one({"id": subgroup_id}, {"$set": {"leader_id": None}})
    
    return {"message": "Member removed from subgroup"}

@api_router.delete("/subgroups/{subgroup_id}")
async def delete_subgroup(subgroup_id: str, current_user: dict = Depends(get_current_user)):
    if not _check_admin_or_tech_support(current_user):
        raise HTTPException(status_code=403, detail="Only admins or Technical Support can manage subgroups")
    
    result = await db.subgroups.delete_one({"id": subgroup_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Subgroup not found")
    return {"message": "Subgroup deleted"}

# ============== JIT TEMPORARY LEADER ==============

def _can_assign_temp_leader(current_user: dict, subgroup: dict) -> bool:
    roles = current_user.get("roles", [])
    if "admin" in roles or "super_admin" in roles:
        return True
    if current_user.get("division") == "Technical Support":
        return True
    if subgroup.get("leader_id") == current_user.get("id"):
        return True
    return False

@api_router.post("/subgroups/{subgroup_id}/temp-leader")
async def assign_temp_leader(subgroup_id: str, body: dict, current_user: dict = Depends(get_current_user)):
    sg = await db.subgroups.find_one({"id": subgroup_id}, {"_id": 0})
    if not sg:
        raise HTTPException(status_code=404, detail="Subgroup not found")
    
    if not _can_assign_temp_leader(current_user, sg):
        raise HTTPException(status_code=403, detail="Only the current leader, admins, or super admins can assign a temporary leader")
    
    temp_leader_id = body.get("temp_leader_id")
    duration_hours = body.get("duration_hours")
    if not temp_leader_id or not duration_hours:
        raise HTTPException(status_code=400, detail="temp_leader_id and duration_hours are required")
    
    if duration_hours < 1:
        raise HTTPException(status_code=400, detail="Duration must be at least 1 hour")
    
    if temp_leader_id not in sg.get("member_user_ids", []):
        raise HTTPException(status_code=400, detail="Temporary leader must be a member of the subgroup")
    
    if temp_leader_id == sg.get("leader_id"):
        raise HTTPException(status_code=400, detail="Cannot assign the current leader as temporary leader")
    
    now = datetime.now(timezone.utc)
    end_time = now + timedelta(hours=duration_hours)
    
    await db.subgroups.update_one({"id": subgroup_id}, {"$set": {
        "temp_leader_id": temp_leader_id,
        "temp_leader_start": now.isoformat(),
        "temp_leader_end": end_time.isoformat(),
        "temp_leader_assigned_by": current_user.get("id"),
        "updated_at": now.isoformat(),
    }})
    
    return {
        "message": "Temporary leader assigned",
        "temp_leader_id": temp_leader_id,
        "temp_leader_start": now.isoformat(),
        "temp_leader_end": end_time.isoformat(),
        "duration_hours": duration_hours,
    }

@api_router.delete("/subgroups/{subgroup_id}/temp-leader")
async def revoke_temp_leader(subgroup_id: str, current_user: dict = Depends(get_current_user)):
    sg = await db.subgroups.find_one({"id": subgroup_id}, {"_id": 0})
    if not sg:
        raise HTTPException(status_code=404, detail="Subgroup not found")
    
    if not _can_assign_temp_leader(current_user, sg):
        raise HTTPException(status_code=403, detail="Only the current leader, admins, or super admins can revoke a temporary leader")
    
    await db.subgroups.update_one({"id": subgroup_id}, {"$set": {
        "temp_leader_id": None,
        "temp_leader_start": None,
        "temp_leader_end": None,
        "temp_leader_assigned_by": None,
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }})
    
    return {"message": "Temporary leader revoked"}

# ============== NOTIFICATIONS ROUTES ==============
# Moved to routers/notifications.py

# ============== GROUP MESSAGING ROUTES ==============

@api_router.post("/messages/group-conversation")
async def create_group_conversation(body: dict, current_user: dict = Depends(get_current_user)):
    target_type = body.get("target_type")  # 'division', 'subgroup', 'individual'
    target_id = body.get("target_id")
    target_name = body.get("target_name", "")
    
    participant_ids = [current_user["id"]]
    conv_name = target_name
    
    if target_type == "division":
        users = await db.users.find({"division": target_name}, {"_id": 0, "id": 1}).to_list(500)
        participant_ids += [u["id"] for u in users if u["id"] != current_user["id"]]
        conv_name = f"Division: {target_name}"
    elif target_type == "subgroup":
        sg = await db.subgroups.find_one({"id": target_id}, {"_id": 0})
        if sg:
            participant_ids += [uid for uid in sg.get("member_user_ids", []) if uid != current_user["id"]]
            conv_name = f"Subgroup: {sg.get('name', '')}"
    elif target_type == "individual":
        participant_ids.append(target_id)
        other = await db.users.find_one({"id": target_id}, {"_id": 0, "full_name": 1})
        conv_name = other.get("full_name", "") if other else ""
    
    participant_ids = list(set(participant_ids))
    
    # Check for existing group conversation with same target
    if target_type in ["division", "subgroup"]:
        existing = await db.conversations.find_one({
            "target_type": target_type,
            "target_id": target_id or target_name,
        }, {"_id": 0})
        if existing:
            participants = await db.users.find(
                {"id": {"$in": existing["participant_ids"]}},
                {"_id": 0, "id": 1, "full_name": 1, "email": 1}
            ).to_list(500)
            existing["participants"] = participants
            return existing
    elif target_type == "individual":
        existing = await db.conversations.find_one({
            "type": "direct",
            "participant_ids": {"$all": participant_ids, "$size": 2}
        }, {"_id": 0})
        if existing:
            participants = await db.users.find(
                {"id": {"$in": existing["participant_ids"]}},
                {"_id": 0, "id": 1, "full_name": 1, "email": 1}
            ).to_list(100)
            existing["participants"] = participants
            return existing
    
    now = datetime.now(timezone.utc).isoformat()
    conv = {
        "id": generate_uuid(),
        "type": "group" if target_type in ["division", "subgroup"] else "direct",
        "name": conv_name,
        "participant_ids": participant_ids,
        "target_type": target_type,
        "target_id": target_id or target_name,
        "created_by": current_user["id"],
        "last_message_at": now,
        "unread_count": {},
        "is_archived": False,
        "created_at": now,
        "updated_at": now,
    }
    await db.conversations.insert_one({**conv})
    
    participants = await db.users.find(
        {"id": {"$in": participant_ids}},
        {"_id": 0, "id": 1, "full_name": 1, "email": 1}
    ).to_list(500)
    conv["participants"] = participants
    
    return conv

# Helper to create notifications from messages
async def _create_message_notifications(conversation, sender_id, sender_name, content):
    for pid in conversation.get("participant_ids", []):
        if pid != sender_id:
            notif = {
                "id": generate_uuid(),
                "user_id": pid,
                "type": "new_message",
                "title": f"New message from {sender_name}",
                "message": content[:100] + ("..." if len(content) > 100 else ""),
                "reference_id": conversation["id"],
                "reference_type": "conversation",
                "is_read": False,
                "created_at": datetime.now(timezone.utc).isoformat(),
            }
            await db.notifications.insert_one({**notif})

# ============== DASHBOARD STATS ==============
# Moved to routers/reports.py

# ============== APP SETUP ==============

# Include extracted routers
from routers.reports import router as reports_router
from routers.notifications import router as notifications_router
from routers.auth import router as auth_router
from routers.applications import router as applications_router
from routers.users import router as users_router
app.include_router(reports_router)
app.include_router(notifications_router)
app.include_router(auth_router)
app.include_router(applications_router)
app.include_router(users_router)

app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
