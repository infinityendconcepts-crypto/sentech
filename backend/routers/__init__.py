"""Shared dependencies for all routers - db, auth, helpers"""
from fastapi import Depends, HTTPException, status
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import StreamingResponse
from motor.motor_asyncio import AsyncIOMotorClient
from passlib.context import CryptContext
from jose import JWTError, jwt
from datetime import datetime, timezone, timedelta
from typing import Optional, Dict, Any, List
from pydantic import BaseModel, Field, ConfigDict
import uuid
import os
import io
import logging
import asyncio
from dotenv import load_dotenv
from pathlib import Path

ROOT_DIR = Path(__file__).parent.parent
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
security = HTTPBearer()

SECRET_KEY = os.getenv("JWT_SECRET_KEY", "your-secret-key-change-in-production-min-32-chars")
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 1440

MICROSOFT_CLIENT_ID = os.getenv("MICROSOFT_CLIENT_ID")
MICROSOFT_TENANT_ID = os.getenv("MICROSOFT_TENANT_ID")
MICROSOFT_CLIENT_SECRET = os.getenv("MICROSOFT_CLIENT_SECRET")
MICROSOFT_AUTHORITY = f"https://login.microsoftonline.com/{MICROSOFT_TENANT_ID}"
MICROSOFT_SCOPE = ["User.Read"]
FRONTEND_URL = os.getenv("REACT_APP_BACKEND_URL", "http://localhost:3000").replace("/api", "")

logger = logging.getLogger(__name__)

def generate_uuid():
    return str(uuid.uuid4())

def current_time():
    return datetime.now(timezone.utc)

def get_password_hash(password: str):
    return pwd_context.hash(password)

def verify_password(plain_password: str, hashed_password: str):
    try:
        return pwd_context.verify(plain_password, hashed_password)
    except Exception:
        return False

def create_access_token(data: dict):
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        payload = jwt.decode(credentials.credentials, SECRET_KEY, algorithms=[ALGORITHM])
        user_id: str = payload.get("sub")
        if user_id is None:
            raise HTTPException(status_code=401, detail="Invalid token")
    except JWTError:
        raise HTTPException(status_code=401, detail="Invalid token")
    user = await db.users.find_one({"id": user_id}, {"_id": 0})
    if user is None:
        raise HTTPException(status_code=401, detail="User not found")
    return user

def is_admin_user(user: dict) -> bool:
    roles = user.get("roles", [])
    return "admin" in roles or "super_admin" in roles

async def send_email_notification(to_email: str, subject: str, body_text: str, link: str = ""):
    smtp_host = os.getenv("SMTP_HOST")
    smtp_port = int(os.getenv("SMTP_PORT", "465"))
    smtp_user = os.getenv("SMTP_USERNAME")
    smtp_pass = os.getenv("SMTP_PASSWORD")
    from_email = os.getenv("SMTP_FROM_EMAIL", "hr@sentechportal.com")
    from_name = os.getenv("SMTP_FROM_NAME", "Sentech Bursary")
    if not smtp_host or not smtp_user:
        logger.info(f"[EMAIL SKIP] No SMTP config. To: {to_email} | Subject: {subject}")
        return
    if not to_email:
        return
    try:
        import smtplib
        from email.mime.text import MIMEText
        from email.mime.multipart import MIMEMultipart

        link_html = ""
        if link:
            link_html = f'<p style="margin:20px 0;"><a href="{link}" style="background:#0056B3;color:#fff;padding:10px 24px;border-radius:6px;text-decoration:none;font-weight:bold;">View in Sentech Portal</a></p>'

        html = f"""
        <div style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto;border:1px solid #e2e8f0;border-radius:8px;overflow:hidden;">
          <div style="background:#0056B3;padding:16px 24px;">
            <h2 style="color:#fff;margin:0;font-size:18px;">Sentech Bursary System</h2>
          </div>
          <div style="padding:24px;">
            <h3 style="color:#1e293b;margin:0 0 12px;">{subject}</h3>
            <div style="color:#475569;font-size:14px;line-height:1.6;">{body_text}</div>
            {link_html}
          </div>
          <div style="background:#f8fafc;padding:12px 24px;font-size:12px;color:#94a3b8;border-top:1px solid #e2e8f0;">
            This is an automated notification from the Sentech Bursary Management System.
          </div>
        </div>
        """

        msg = MIMEMultipart()
        msg["From"] = f"{from_name} <{from_email}>"
        msg["To"] = to_email
        msg["Subject"] = subject
        msg.attach(MIMEText(html, "html"))

        if smtp_port == 465:
            with smtplib.SMTP_SSL(smtp_host, smtp_port) as server:
                server.login(smtp_user, smtp_pass)
                server.sendmail(from_email, to_email, msg.as_string())
        else:
            with smtplib.SMTP(smtp_host, smtp_port) as server:
                server.starttls()
                server.login(smtp_user, smtp_pass)
                server.sendmail(from_email, to_email, msg.as_string())
        logger.info(f"[EMAIL SENT] To: {to_email} | Subject: {subject}")
    except Exception as e:
        logger.error(f"[EMAIL FAILED] To: {to_email} | Error: {e}")


def get_app_url():
    return os.getenv("APP_URL", "https://rbac-forms-dev.preview.emergentagent.com")

async def notify_and_email(user_id: str, title: str, message: str, ref_id: str = "", ref_type: str = "", link_path: str = ""):
    """Create an in-app notification AND send an email to the user."""
    now = datetime.now(timezone.utc).isoformat()
    notif = {
        "id": generate_uuid(), "user_id": user_id, "type": "status_change",
        "title": title, "message": message,
        "reference_id": ref_id, "reference_type": ref_type,
        "is_read": False, "created_at": now,
    }
    await db.notifications.insert_one({**notif})
    user = await db.users.find_one({"id": user_id}, {"_id": 0, "email": 1})
    if user and user.get("email"):
        link = f"{get_app_url()}{link_path}" if link_path else f"{get_app_url()}/notifications"
        asyncio.create_task(send_email_notification(user["email"], title, f"<p>{message}</p>", link))

async def notify_admins_and_heads(title: str, message: str, ref_id: str = "", ref_type: str = "", link_path: str = "", exclude_user_id: str = ""):
    """Send notification + email to all admins, super_admins, and heads."""
    recipients = await db.users.find(
        {"$or": [{"roles": {"$in": ["admin", "super_admin"]}}, {"is_head": True}]},
        {"_id": 0, "id": 1, "email": 1}
    ).to_list(500)
    for r in recipients:
        if r["id"] == exclude_user_id:
            continue
        await notify_and_email(r["id"], title, message, ref_id, ref_type, link_path)

def generate_excel(data: list, title: str = "Export") -> io.BytesIO:
    """Generate an Excel file from a list of dicts"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    if not data:
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    headers = list(data[0].keys())
    header_fill = PatternFill(start_color="0056B3", end_color="0056B3", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header.replace("_", " ").title())
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    alt_fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, header in enumerate(headers, 1):
            val = row_data.get(header, "")
            if isinstance(val, list):
                val = ", ".join(str(v) for v in val)
            elif isinstance(val, dict):
                val = str(val)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    for col_idx, header in enumerate(headers, 1):
        max_len = max(len(str(header)), max((len(str(row_data.get(header, ""))) for row_data in data), default=0))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_pdf(data: list, title: str = "Export") -> io.BytesIO:
    """Generate a PDF file from a list of dicts"""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4, rightMargin=0.5*inch, leftMargin=0.5*inch,
                            topMargin=0.5*inch, bottomMargin=0.5*inch, title=title)
    elements = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle('title', parent=styles['Heading1'], fontSize=16,
                                  textColor=colors.HexColor('#0056B3'), spaceAfter=12, alignment=1)
    elements.append(Paragraph(title, title_style))
    elements.append(Spacer(1, 0.1*inch))
    meta = f"<font size=9>Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')} &nbsp;&nbsp; Total: {len(data)} records</font>"
    elements.append(Paragraph(meta, styles['Normal']))
    elements.append(Spacer(1, 0.15*inch))

    if not data:
        elements.append(Paragraph("No data available.", styles['Normal']))
        doc.build(elements)
        output.seek(0)
        return output

    headers = list(data[0].keys())
    display_headers = [h.replace("_", " ").title() for h in headers]
    table_data = [display_headers]
    for row in data:
        r = []
        for h in headers:
            val = row.get(h, "")
            if isinstance(val, list):
                val = ", ".join(str(v) for v in val)
            elif isinstance(val, dict):
                val = str(val)
            r.append(str(val)[:80])
        table_data.append(r)

    col_width = (A4[0] - inch) / len(headers)
    table = Table(table_data, colWidths=[col_width]*len(headers), repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0056B3')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#E8F0FE')]),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('PADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(table)
    doc.build(elements)
    output.seek(0)
    return output
